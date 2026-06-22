#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
TikTok live-commerce report generator.

Pokemon dashboard version plus:
- package metrics for awaiting-shipment orders
- new / returning customer metrics
- pre-shipment and post-shipment refund split
"""

from __future__ import annotations

import csv
import re
import sys
from collections import Counter, defaultdict
from copy import copy
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple
from zoneinfo import ZoneInfo

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


SCRIPT_VERSION = "tiktok_report_generator_v20260617_004_view_format_followup"
SOURCE_DIR_NAME = "源数据文件"
OUTPUT_PREFIX = "TikTok_数据汇报"

LIVE_HEADER_KEY = "直播间信息"
ORDER_ID_HEADER = "Order ID"
LIVE_FILE_HINTS = ("creator-live-performance", "live-performance", "live")
ORDER_FILE_HINTS = ("all order", "allorder", "order")
PACKAGE_FILE_HINTS = ("发货表", "package")
MAX_ORDERS_PER_PACKAGE = 20
TIKTOK_EXPORT_TZ = ZoneInfo("America/Los_Angeles")
REPORT_TZ = ZoneInfo("Asia/Shanghai")
NUMBER_FORMAT = "#,##0.0"
CURRENCY_FORMAT = "$#,##0.0"
PERCENT_FORMAT = "0.0%"
MIN_LIVE_SECONDS = Decimal("1200")


def default_data_dir(script_dir: Path) -> Path:
    return script_dir.parents[1] / "data" / "汇报"

PALETTE = {
    "navy": "1B2A41",
    "blue": "2A75BB",
    "red": "E3350D",
    "yellow": "FFCB05",
    "cream": "FFF7D6",
    "sky": "D9ECFF",
    "orange": "F28C28",
    "green": "3BB273",
    "gray": "6B7280",
    "line": "D7DEE8",
    "white": "FFFFFF",
    "black": "111827",
    "band": "F8FBFF",
}


@dataclass
class SourceBundle:
    live_files: List[Path]
    order_files: List[Path]
    package_files: List[Path]
    selected_order_file: Optional[Path]
    selected_package_file: Optional[Path]


@dataclass
class Order:
    order_id: str
    order_date: Optional[date]
    created_at: Optional[datetime]
    status: str
    substatus: str
    amount: Decimal
    refund: Decimal
    quantity: Decimal
    sku_rows: int
    customer_key: str
    customer_name: str
    recipient: str
    phone: str
    country: str
    state: str
    city: str
    zipcode: str
    address1: str
    address2: str
    delivery_instruction: str
    tracking_id: str
    shipped_at: Optional[datetime]
    delivered_at: Optional[datetime]
    rts_at: Optional[datetime]
    source_file: str
    created_time_raw: str
    created_time_report: str

    @property
    def is_pending_shipment(self) -> bool:
        text = f"{self.status} {self.substatus}".lower()
        return "to ship" in text or "awaiting shipment" in text

    @property
    def is_delivered(self) -> bool:
        text = f"{self.status} {self.substatus}".lower()
        return "delivered" in text or self.delivered_at is not None

    @property
    def is_canceled(self) -> bool:
        text = f"{self.status} {self.substatus}".lower()
        return "cancel" in text

    @property
    def is_shipped_package(self) -> bool:
        text = f"{self.status} {self.substatus}".lower()
        return bool(self.tracking_id) and (
            "shipped" in text
            or "in transit" in text
            or "delivered" in text
            or self.shipped_at is not None
            or self.delivered_at is not None
            or self.rts_at is not None
        )

    @property
    def is_shipped_before_refund(self) -> bool:
        return self.shipped_at is not None or self.delivered_at is not None


@dataclass
class Package:
    package_id: str
    package_date: date
    package_rule: str
    tracking_id: str
    order_ids: List[str]
    customer_name: str
    recipient: str
    zipcode: str
    address: str
    amount: Decimal


@dataclass
class PeriodSummary:
    key: str
    label: str
    start_date: date
    end_date: date
    live_sessions: int = 0
    live_seconds: Decimal = Decimal("0")
    live_gmv: Decimal = Decimal("0")
    live_orders: Decimal = Decimal("0")
    live_units: Decimal = Decimal("0")
    views: Decimal = Decimal("0")
    exposure: Decimal = Decimal("0")
    product_clicks: Decimal = Decimal("0")
    new_followers: Decimal = Decimal("0")
    likes: Decimal = Decimal("0")
    comments: Decimal = Decimal("0")
    shares: Decimal = Decimal("0")
    ad_cost: Decimal = Decimal("0")
    ad_gmv: Decimal = Decimal("0")
    order_count: int = 0
    units: Decimal = Decimal("0")
    order_amount: Decimal = Decimal("0")
    refund_amount: Decimal = Decimal("0")
    pre_ship_refund_amount: Decimal = Decimal("0")
    post_ship_refund_amount: Decimal = Decimal("0")
    pre_ship_refund_orders: int = 0
    post_ship_refund_orders: int = 0
    pending_orders: int = 0
    delivered_orders: int = 0
    canceled_orders: int = 0
    packages: int = 0
    package_amount: Decimal = Decimal("0")
    new_customers: int = 0
    returning_customers: int = 0
    active_customers: int = 0
    returning_orders: int = 0
    returning_sales: Decimal = Decimal("0")

    @property
    def net_sales(self) -> Decimal:
        return self.order_amount - self.refund_amount

    @property
    def aov(self) -> Decimal:
        return div_decimal(self.order_amount, Decimal(self.order_count))

    @property
    def refund_rate(self) -> Decimal:
        return div_decimal(self.refund_amount, self.order_amount)

    @property
    def package_order_ratio(self) -> Decimal:
        return div_decimal(Decimal(self.order_count), Decimal(self.packages))

    @property
    def avg_package_value(self) -> Decimal:
        return div_decimal(self.order_amount, Decimal(self.packages))

    @property
    def avg_gmv_per_live(self) -> Decimal:
        return div_decimal(self.live_gmv, Decimal(self.live_sessions))

    @property
    def live_hours(self) -> Decimal:
        return div_decimal(self.live_seconds, Decimal("3600"))

    @property
    def gmv_per_live_hour(self) -> Decimal:
        return div_decimal(self.live_gmv, self.live_hours)

    @property
    def returning_customer_rate(self) -> Decimal:
        return div_decimal(Decimal(self.returning_customers), Decimal(self.active_customers))

    @property
    def returning_order_rate(self) -> Decimal:
        return div_decimal(Decimal(self.returning_orders), Decimal(self.order_count))

    @property
    def view_order_rate(self) -> Decimal:
        return div_decimal(self.live_orders, self.views)

    @property
    def click_order_rate(self) -> Decimal:
        return div_decimal(self.live_orders, self.product_clicks)

    @property
    def engagement_rate(self) -> Decimal:
        return div_decimal(self.likes + self.comments + self.shares, self.views)

    @property
    def roas(self) -> Decimal:
        return div_decimal(self.ad_gmv, self.ad_cost)


def clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, Decimal):
        return str(value)
    return str(value).strip()


def norm_key(value: Any) -> str:
    return re.sub(r"\s+", " ", clean(value)).strip().lower()


def to_decimal(value: Any) -> Decimal:
    text = clean(value).replace("$", "").replace(",", "").replace("%", "").strip()
    if not text or text.lower() in {"nan", "none", "-"}:
        return Decimal("0")
    try:
        return Decimal(text)
    except InvalidOperation:
        return Decimal("0")


def parse_live_duration_seconds(value: Any) -> Decimal:
    if isinstance(value, timedelta):
        return Decimal(str(value.total_seconds()))

    text = clean(value)
    if not text:
        return Decimal("0")

    if ":" in text:
        parts = text.split(":")
        try:
            numbers = [Decimal(part) for part in parts]
        except InvalidOperation:
            return Decimal("0")
        if len(numbers) == 3:
            return numbers[0] * Decimal("3600") + numbers[1] * Decimal("60") + numbers[2]
        if len(numbers) == 2:
            return numbers[0] * Decimal("60") + numbers[1]

    duration = to_decimal(value)
    if Decimal("0") < duration < Decimal("1"):
        return duration * Decimal("86400")
    return duration


def div_decimal(numerator: Decimal, denominator: Decimal) -> Decimal:
    return Decimal("0") if denominator == 0 else numerator / denominator


def as_float(value: Decimal | int | float) -> float:
    return float(value)


def parse_datetime(value: Any) -> Optional[datetime]:
    if isinstance(value, datetime):
        return value
    text = clean(value).replace("\t", "").strip()
    if not text:
        return None
    for fmt in (
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d %H:%M:%S",
        "%m/%d/%Y %I:%M:%S %p",
        "%m/%d/%Y %H:%M:%S",
        "%Y/%m/%d %H:%M",
        "%Y/%m/%d %H:%M:%S",
    ):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def tiktok_time_to_report_time(value: Any) -> Optional[datetime]:
    parsed = parse_datetime(value)
    if parsed is None:
        return None
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=TIKTOK_EXPORT_TZ)
    return parsed.astimezone(REPORT_TZ).replace(tzinfo=None)


def unique_headers(headers: Sequence[Any]) -> List[str]:
    counts: Dict[str, int] = defaultdict(int)
    result: List[str] = []
    for raw in headers:
        name = clean(raw) or "Unnamed"
        counts[name] += 1
        result.append(name if counts[name] == 1 else f"{name}_{counts[name]}")
    return result


def find_sources(base_dir: Path) -> SourceBundle:
    source_dir = base_dir / SOURCE_DIR_NAME
    if not source_dir.exists():
        raise FileNotFoundError(f"找不到源数据文件夹：{source_dir}")
    live_files: List[Path] = []
    order_files: List[Path] = []
    package_files: List[Path] = []
    for path in sorted(source_dir.iterdir()):
        if path.name.startswith("~$") or path.suffix.lower() not in {".xlsx", ".csv"}:
            continue
        lower = path.name.lower()
        if any(hint in lower for hint in LIVE_FILE_HINTS):
            live_files.append(path)
        elif any(hint in lower for hint in PACKAGE_FILE_HINTS):
            package_files.append(path)
        elif any(hint in lower for hint in ORDER_FILE_HINTS):
            order_files.append(path)
    selected = max(order_files, key=lambda p: p.stat().st_mtime, default=None)
    selected_package = max(package_files, key=lambda p: p.stat().st_mtime, default=None)
    return SourceBundle(live_files, order_files, package_files, selected, selected_package)


def read_xlsx_rows(path: Path) -> List[List[Any]]:
    wb = load_workbook(path, read_only=False, data_only=True)
    return [list(row) for row in wb.active.iter_rows(values_only=True)]


def read_csv_rows(path: Path) -> List[List[Any]]:
    for encoding in ("utf-8-sig", "utf-8", "gb18030", "gbk"):
        try:
            with path.open("r", encoding=encoding, newline="") as handle:
                return [row for row in csv.reader(handle)]
        except UnicodeDecodeError:
            continue
    raise ValueError(f"无法读取 CSV 编码：{path}")


def read_rows(path: Path) -> List[List[Any]]:
    return read_csv_rows(path) if path.suffix.lower() == ".csv" else read_xlsx_rows(path)


def rows_to_records(rows: List[List[Any]], header_index: int, source_file: Path) -> List[Dict[str, Any]]:
    headers = unique_headers(rows[header_index])
    records: List[Dict[str, Any]] = []
    for row in rows[header_index + 1 :]:
        if not any(clean(cell) for cell in row):
            continue
        record = {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}
        record["Source File"] = source_file.name
        records.append(record)
    return records


def read_live_records(files: Iterable[Path]) -> List[Dict[str, Any]]:
    records: List[Dict[str, Any]] = []
    seen = set()
    for path in files:
        rows = read_rows(path)
        header_index = next((i for i, row in enumerate(rows[:30]) if any(clean(cell) == LIVE_HEADER_KEY for cell in row)), None)
        if header_index is None:
            continue
        for record in rows_to_records(rows, header_index, path):
            live_duration_seconds = parse_live_duration_seconds(record.get("直播时长"))
            if live_duration_seconds < MIN_LIVE_SECONDS:
                continue

            started_at_raw = parse_datetime(record.get("直播开始时间"))
            started_at_report = tiktok_time_to_report_time(record.get("直播开始时间"))
            if not started_at_raw or not started_at_report:
                continue
            key = (clean(record.get("直播间信息")), started_at_raw.strftime("%Y-%m-%d %H:%M"))
            if key in seen:
                continue
            seen.add(key)
            record["直播日期"] = started_at_report.date().isoformat()
            record["直播开始时间_dt"] = started_at_report
            record["直播开始时间_原始"] = started_at_raw.strftime("%Y-%m-%d %H:%M:%S")
            record["直播开始时间_上海"] = started_at_report.strftime("%Y-%m-%d %H:%M:%S")
            records.append(record)
    return sorted(records, key=lambda item: item["直播开始时间_dt"])


def read_order_records(order_file: Optional[Path]) -> List[Dict[str, Any]]:
    if order_file is None:
        return []
    rows = read_rows(order_file)
    header_index = next((i for i, row in enumerate(rows[:10]) if any(clean(cell) == ORDER_ID_HEADER for cell in row)), None)
    if header_index is None:
        return []
    records: List[Dict[str, Any]] = []
    seen = set()
    for record in rows_to_records(rows, header_index, order_file):
        order_id = clean(record.get("Order ID"))
        if not order_id or order_id.lower().startswith("platform unique order id"):
            continue
        key = (order_id, clean(record.get("SKU ID")), clean(record.get("Product Name")), clean(record.get("Created Time")))
        if key in seen:
            continue
        seen.add(key)
        created_at_raw = parse_datetime(record.get("Created Time") or record.get("Paid Time"))
        created_at_report = tiktok_time_to_report_time(record.get("Created Time") or record.get("Paid Time"))
        record["订单日期"] = created_at_report.date().isoformat() if created_at_report else ""
        record["Created Time_dt"] = created_at_report
        record["Created Time_raw_dt"] = created_at_raw
        record["Created Time_原始"] = created_at_raw.strftime("%Y-%m-%d %H:%M:%S") if created_at_raw else ""
        record["Created Time_上海"] = created_at_report.strftime("%Y-%m-%d %H:%M:%S") if created_at_report else ""
        records.append(record)
    return records


def sum_field(records: Iterable[Dict[str, Any]], field: str) -> Decimal:
    return sum((to_decimal(record.get(field)) for record in records), Decimal("0"))


def customer_key_from_row(row: Dict[str, Any]) -> Tuple[str, str]:
    username = clean(row.get("Buyer Username"))
    nickname = clean(row.get("Buyer Nickname"))
    recipient = clean(row.get("Recipient"))
    phone = clean(row.get("Phone #"))
    if username:
        return f"username:{username.lower()}", nickname or username
    if nickname:
        return f"nickname:{nickname.lower()}", nickname
    fallback = f"{recipient}|{phone}".lower()
    return f"recipient:{fallback}", recipient or phone or "Unknown"


def order_level_records(order_records: List[Dict[str, Any]]) -> List[Order]:
    grouped: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for record in order_records:
        grouped[clean(record.get("Order ID"))].append(record)

    orders: List[Order] = []
    for order_id, rows in grouped.items():
        first = rows[0]
        created_at = first.get("Created Time_dt") or parse_datetime(first.get("Created Time") or first.get("Paid Time"))
        amount = next((to_decimal(row.get("Order Amount")) for row in rows if to_decimal(row.get("Order Amount"))), Decimal("0"))
        refund = next((to_decimal(row.get("Order Refund Amount")) for row in rows if to_decimal(row.get("Order Refund Amount"))), Decimal("0"))
        customer_key, customer_name = customer_key_from_row(first)
        orders.append(
            Order(
                order_id=order_id,
                order_date=created_at.date() if created_at else None,
                created_at=created_at,
                status=clean(first.get("Order Status")) or "Unknown",
                substatus=clean(first.get("Order Substatus")) or "Unknown",
                amount=amount,
                refund=refund,
                quantity=sum_field(rows, "Quantity"),
                sku_rows=len(rows),
                customer_key=customer_key,
                customer_name=customer_name,
                recipient=clean(first.get("Recipient")),
                phone=clean(first.get("Phone #")),
                country=clean(first.get("Country")),
                state=clean(first.get("State")),
                city=clean(first.get("City")),
                zipcode=clean(first.get("Zipcode")),
                address1=clean(first.get("Address Line 1")),
                address2=clean(first.get("Address Line 2")),
                delivery_instruction=clean(first.get("Delivery Instruction")),
                tracking_id=clean(first.get("Tracking ID")),
                shipped_at=parse_datetime(first.get("Shipped Time")),
                delivered_at=parse_datetime(first.get("Delivered Time")),
                rts_at=parse_datetime(first.get("RTS Time")),
                source_file=clean(first.get("Source File")),
                created_time_raw=clean(first.get("Created Time_原始")),
                created_time_report=clean(first.get("Created Time_上海")),
            )
        )
    return sorted(orders, key=lambda item: item.created_at or datetime.min)


def package_group_key(order: Order) -> Tuple[str, str, str, str, str, str]:
    return (
        norm_key(order.recipient),
        norm_key(order.zipcode),
        norm_key(order.country),
        norm_key(order.state),
        norm_key(order.city),
        norm_key(f"{order.address1} {order.address2} {order.delivery_instruction}"),
    )


def build_packages(orders: List[Order]) -> List[Package]:
    shipped_by_tracking: Dict[str, List[Order]] = defaultdict(list)
    by_day_group: Dict[Tuple[date, Tuple[str, str, str, str, str, str]], List[Order]] = defaultdict(list)
    for order in orders:
        if order.is_canceled:
            continue
        if order.is_shipped_package:
            shipped_by_tracking[order.tracking_id].append(order)
        elif order.order_date and order.is_pending_shipment:
            by_day_group[(order.order_date, package_group_key(order))].append(order)

    packages: List[Package] = []
    for tracking_id, group_orders in sorted(shipped_by_tracking.items(), key=lambda item: item[0]):
        group_orders = sorted(group_orders, key=lambda item: item.created_at or datetime.min)
        first = group_orders[0]
        pkg_date = first.shipped_at or first.delivered_at or first.rts_at or first.created_at
        if pkg_date is None:
            continue
        packages.append(
            Package(
                package_id=f"TRK-{tracking_id}",
                package_date=pkg_date.date(),
                package_rule="已发货-快递单号",
                tracking_id=tracking_id,
                order_ids=[order.order_id for order in group_orders],
                customer_name=first.customer_name,
                recipient=first.recipient,
                zipcode=first.zipcode,
                address=", ".join(part for part in [first.address1, first.address2, first.city, first.state, first.country] if part),
                amount=sum((order.amount for order in group_orders), Decimal("0")),
            )
        )

    for (pkg_date, _group_key), group_orders in sorted(by_day_group.items(), key=lambda item: (item[0][0], item[0][1])):
        group_orders = sorted(group_orders, key=lambda item: item.created_at or datetime.min)
        for idx in range(0, len(group_orders), MAX_ORDERS_PER_PACKAGE):
            chunk = group_orders[idx : idx + MAX_ORDERS_PER_PACKAGE]
            first = chunk[0]
            seq = (idx // MAX_ORDERS_PER_PACKAGE) + 1
            packages.append(
                Package(
                    package_id=f"{pkg_date:%Y%m%d}-{len(packages)+1:04d}-{seq}",
                    package_date=pkg_date,
                    package_rule="待发货-同地址",
                    tracking_id="",
                    order_ids=[order.order_id for order in chunk],
                    customer_name=first.customer_name,
                    recipient=first.recipient,
                    zipcode=first.zipcode,
                    address=", ".join(part for part in [first.address1, first.address2, first.city, first.state, first.country] if part),
                    amount=sum((order.amount for order in chunk), Decimal("0")),
                )
            )
    return packages


def is_awaiting_shipment_order(order: Order) -> bool:
    return order.substatus.lower() == "awaiting shipment"


def build_daily_awaiting_shipment_packages(orders: List[Order]) -> List[Package]:
    by_day_group: Dict[Tuple[date, Tuple[str, str, str, str, str, str]], List[Order]] = defaultdict(list)
    for order in orders:
        if not order.order_date or order.is_canceled or not is_awaiting_shipment_order(order):
            continue
        by_day_group[(order.order_date, package_group_key(order))].append(order)

    packages: List[Package] = []
    for (pkg_date, _group_key), group_orders in sorted(by_day_group.items(), key=lambda item: (item[0][0], item[0][1])):
        group_orders = sorted(group_orders, key=lambda item: item.created_at or datetime.min)
        for idx in range(0, len(group_orders), MAX_ORDERS_PER_PACKAGE):
            chunk = group_orders[idx : idx + MAX_ORDERS_PER_PACKAGE]
            if not chunk:
                continue
            first = chunk[0]
            seq = (idx // MAX_ORDERS_PER_PACKAGE) + 1
            packages.append(
                Package(
                    package_id=f"DAILY-AWAITING-{pkg_date:%Y%m%d}-{len(packages)+1:04d}-{seq}",
                    package_date=pkg_date,
                    package_rule="日报待发货-同收货信息",
                    tracking_id="",
                    order_ids=[order.order_id for order in chunk],
                    customer_name=first.customer_name,
                    recipient=first.recipient,
                    zipcode=first.zipcode,
                    address=", ".join(part for part in [first.address1, first.address2, first.city, first.state, first.country] if part),
                    amount=sum((order.amount for order in chunk), Decimal("0")),
                )
            )
    return packages


def add_order_snapshot_packages_to_summary(
    summary: Dict[str, PeriodSummary],
    orders: List[Order],
    grain: str,
) -> None:
    grouped: Dict[Tuple[str, Tuple[str, str, str, str, str, str]], List[Order]] = defaultdict(list)
    for order in orders:
        if not order.order_date or order.is_canceled:
            continue
        key, _label, _start, _end = period_key(order.order_date, grain)
        grouped[(key, package_group_key(order))].append(order)

    for (key, _group_key), group_orders in grouped.items():
        group_orders = sorted(group_orders, key=lambda item: item.created_at or datetime.min)
        if not group_orders:
            continue
        item = get_period(summary, group_orders[0].order_date, grain)
        for start in range(0, len(group_orders), MAX_ORDERS_PER_PACKAGE):
            chunk = group_orders[start : start + MAX_ORDERS_PER_PACKAGE]
            if not chunk:
                continue
            item.packages += 1
            item.package_amount += sum((order.amount for order in chunk), Decimal("0"))


def infer_report_year(orders: List[Order]) -> int:
    dated_orders = [order.order_date for order in orders if order.order_date]
    if dated_orders:
        return max(dated_orders).year
    return datetime.now().year


def parse_package_sheet_date(sheet_name: str, year: int) -> Optional[date]:
    match = re.search(r"(\d{4})", sheet_name)
    if not match:
        return None
    text = match.group(1)
    month = int(text[:2])
    day = int(text[2:])
    try:
        return date(year, month, day)
    except ValueError:
        return None


def read_packages_from_table(package_file: Optional[Path], year: int) -> List[Package]:
    if package_file is None:
        return []

    # The package workbook can have stale worksheet dimensions; normal mode is
    # more reliable than read-only mode for reading rows past A1.
    wb = load_workbook(package_file, read_only=False, data_only=True)
    packages: List[Package] = []
    seen_ids = set()
    package_header_aliases = {"packageid", "package id", "包裹id", "包裹号", "包裹编号"}
    tracking_header_aliases = {"trackingid", "tracking id", "快递单号", "物流单号", "运单号", "单号"}

    for ws in wb.worksheets:
        package_date = parse_package_sheet_date(ws.title, year)
        if package_date is None:
            continue

        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = next(rows_iter)
        except StopIteration:
            continue

        package_col = None
        tracking_col = None
        for idx, value in enumerate(header):
            header_key = clean(value).strip().lower()
            if header_key in package_header_aliases:
                package_col = idx
            if header_key in tracking_header_aliases:
                tracking_col = idx
        if package_col is None:
            continue

        for row in rows_iter:
            if package_col >= len(row):
                continue
            package_id = clean(row[package_col])
            if not package_id or package_id.lower() in package_header_aliases or package_id in seen_ids:
                continue
            tracking_id = clean(row[tracking_col]) if tracking_col is not None and tracking_col < len(row) else ""
            seen_ids.add(package_id)
            packages.append(
                Package(
                    package_id=package_id,
                    package_date=package_date,
                    package_rule="包裹表-PackageID",
                    tracking_id=tracking_id,
                    order_ids=[],
                    customer_name="",
                    recipient="",
                    zipcode="",
                    address="",
                    amount=Decimal("0"),
                )
            )

    return sorted(packages, key=lambda item: (item.package_date, item.package_id))


def build_tracking_fallback_packages(orders: List[Order], existing_packages: List[Package]) -> List[Package]:
    existing_tracking_ids = {package.tracking_id for package in existing_packages if package.tracking_id}
    latest_package_date = max((package.package_date for package in existing_packages), default=None)

    grouped: Dict[str, List[Order]] = defaultdict(list)
    for order in orders:
        tracking_id = clean(order.tracking_id)
        if not tracking_id or order.is_canceled or not order.is_shipped_package:
            continue
        if tracking_id in existing_tracking_ids:
            continue
        shipped_date = order.shipped_at or order.delivered_at or order.rts_at or order.created_at
        if shipped_date is None:
            continue
        if not existing_tracking_ids and latest_package_date is not None and shipped_date.date() <= latest_package_date:
            continue
        grouped[tracking_id].append(order)

    packages: List[Package] = []
    for tracking_id, group_orders in sorted(grouped.items(), key=lambda item: item[0]):
        group_orders = sorted(group_orders, key=lambda item: item.created_at or datetime.min)
        first = group_orders[0]
        pkg_dt = first.shipped_at or first.delivered_at or first.rts_at or first.created_at
        if pkg_dt is None:
            continue
        packages.append(
            Package(
                package_id=f"TRK-FALLBACK-{tracking_id}",
                package_date=pkg_dt.date(),
                package_rule="订单快递单号-补充包裹",
                tracking_id=tracking_id,
                order_ids=[order.order_id for order in group_orders],
                customer_name=first.customer_name,
                recipient=first.recipient,
                zipcode=first.zipcode,
                address=", ".join(part for part in [first.address1, first.address2, first.city, first.state, first.country] if part),
                amount=sum((order.amount for order in group_orders), Decimal("0")),
            )
        )
    return packages


def period_key(day: date, grain: str) -> Tuple[str, str, date, date]:
    if grain == "day":
        return day.isoformat(), day.strftime("%m-%d"), day, day
    if grain == "week":
        start = day - timedelta(days=day.weekday())
        end = start + timedelta(days=6)
        iso_year, iso_week, _ = day.isocalendar()
        return f"{iso_year}-W{iso_week:02d}", f"W{iso_week:02d}\n{start:%m-%d}~{end:%m-%d}", start, end
    start = day.replace(day=1)
    end = (start.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)
    return start.strftime("%Y-%m"), start.strftime("%Y-%m"), start, end


def get_period(summary: Dict[str, PeriodSummary], day: date, grain: str) -> PeriodSummary:
    key, label, start, end = period_key(day, grain)
    if key not in summary:
        summary[key] = PeriodSummary(key, label, start, end)
    return summary[key]


def customer_first_dates(orders: List[Order]) -> Dict[str, date]:
    result: Dict[str, date] = {}
    for order in orders:
        if not order.order_date:
            continue
        if order.customer_key not in result or order.order_date < result[order.customer_key]:
            result[order.customer_key] = order.order_date
    return result


def customer_period_stats(orders: List[Order], grain: str) -> List[List[Any]]:
    grain_label = {"day": "日", "week": "周", "month": "月"}.get(grain, grain)
    first_dates = customer_first_dates(orders)
    grouped: Dict[Tuple[str, str], Dict[str, Any]] = {}
    for order in orders:
        if not order.order_date:
            continue
        key, label, start, end = period_key(order.order_date, grain)
        group_key = (key, order.customer_key)
        if group_key not in grouped:
            grouped[group_key] = {
                "period_key": key,
                "period": label,
                "start": start,
                "end": end,
                "customer": order.customer_name,
                "customer_key": order.customer_key,
                "orders": 0,
                "sales": Decimal("0"),
                "refund": Decimal("0"),
                "first_date": first_dates.get(order.customer_key),
            }
        item = grouped[group_key]
        item["orders"] += 1
        item["sales"] += order.amount
        item["refund"] += order.refund

    rows = []
    for item in grouped.values():
        first_date = item["first_date"]
        customer_type = "新客" if first_date and item["start"] <= first_date <= item["end"] else "老客"
        rows.append([
            grain_label,
            item["period_key"],
            item["period"],
            item["customer"],
            customer_type,
            item["orders"],
            as_float(item["sales"]),
            as_float(item["refund"]),
            first_date.isoformat() if first_date else "",
        ])
    return sorted(rows, key=lambda row: (row[0], row[1], -row[6], row[3]))


def top_customer_rows_for_period(orders: List[Order], grain: str, latest_period: Optional[PeriodSummary], limit: int = 8) -> List[List[Any]]:
    if latest_period is None:
        return []
    rows = [
        row
        for row in customer_period_stats(orders, grain)
        if row[1] == latest_period.key
    ]
    return sorted(rows, key=lambda row: row[6], reverse=True)[:limit]


def build_period_summaries(
    live_records: List[Dict[str, Any]],
    orders: List[Order],
    packages: List[Package],
    grain: str,
) -> List[PeriodSummary]:
    summary: Dict[str, PeriodSummary] = {}
    first_dates = customer_first_dates(orders)
    customers_by_period: Dict[str, set] = defaultdict(set)
    returning_customers_by_period: Dict[str, set] = defaultdict(set)
    new_customers_by_period: Dict[str, set] = defaultdict(set)

    for record in live_records:
        started_at = record.get("直播开始时间_dt")
        if not isinstance(started_at, datetime):
            continue
        item = get_period(summary, started_at.date(), grain)
        item.live_sessions += 1
        item.live_seconds += to_decimal(record.get("直播时长"))
        item.live_gmv += to_decimal(record.get("归因 GMV"))
        item.live_orders += to_decimal(record.get("归因订单数"))
        item.live_units += to_decimal(record.get("归因成交件数"))
        item.views += to_decimal(record.get("观看次数"))
        item.exposure += to_decimal(record.get("直播曝光次数"))
        item.product_clicks += to_decimal(record.get("商品点击量"))
        item.new_followers += to_decimal(record.get("新粉数"))
        item.likes += to_decimal(record.get("点赞数"))
        item.comments += to_decimal(record.get("评论数"))
        item.shares += to_decimal(record.get("分享次数"))
        item.ad_cost += to_decimal(record.get("广告成本"))
        item.ad_gmv += to_decimal(record.get("广告 GMV"))

    for order in orders:
        if not order.order_date:
            continue
        item = get_period(summary, order.order_date, grain)
        item.order_count += 1
        item.units += order.quantity
        item.order_amount += order.amount
        item.refund_amount += order.refund
        item.pending_orders += 1 if order.is_pending_shipment else 0
        item.delivered_orders += 1 if order.is_delivered else 0
        item.canceled_orders += 1 if order.is_canceled else 0
        if order.refund > 0:
            if order.is_shipped_before_refund:
                item.post_ship_refund_orders += 1
                item.post_ship_refund_amount += order.refund
            else:
                item.pre_ship_refund_orders += 1
                item.pre_ship_refund_amount += order.refund

        first_date = first_dates.get(order.customer_key)
        customers_by_period[item.key].add(order.customer_key)
        if first_date and item.start_date <= first_date <= item.end_date:
            new_customers_by_period[item.key].add(order.customer_key)
        elif first_date and first_date < item.start_date:
            returning_customers_by_period[item.key].add(order.customer_key)
            item.returning_orders += 1
            item.returning_sales += order.amount

    add_order_snapshot_packages_to_summary(summary, orders, grain)

    for key, item in summary.items():
        item.active_customers = len(customers_by_period[key])
        item.new_customers = len(new_customers_by_period[key])
        item.returning_customers = len(returning_customers_by_period[key])

    return sorted(summary.values(), key=lambda item: item.start_date)


def period_rows(periods: List[PeriodSummary]) -> List[List[Any]]:
    return [
        [
            item.key,
            item.label,
            item.start_date.isoformat(),
            item.end_date.isoformat(),
            item.live_sessions,
            as_float(item.avg_gmv_per_live),
            as_float(item.gmv_per_live_hour),
            as_float(item.live_gmv),
            int(item.live_orders),
            int(item.views),
            as_float(item.view_order_rate),
            as_float(item.roas),
            item.order_count,
            as_float(item.order_amount),
            as_float(item.refund_amount),
            as_float(item.refund_rate),
            as_float(item.pre_ship_refund_amount),
            item.pre_ship_refund_orders,
            as_float(item.post_ship_refund_amount),
            item.post_ship_refund_orders,
            item.pending_orders,
            item.packages,
            as_float(item.package_order_ratio),
            as_float(item.avg_package_value),
            item.active_customers,
            item.new_customers,
            item.returning_customers,
            as_float(item.returning_customer_rate),
            item.returning_orders,
            as_float(item.returning_order_rate),
            as_float(item.returning_sales),
        ]
        for item in periods
    ]


def product_ranking(order_records: List[Dict[str, Any]], limit: int = 20) -> List[List[Any]]:
    stats: Dict[str, Dict[str, Decimal | int]] = defaultdict(lambda: {"units": Decimal("0"), "sales": Decimal("0"), "rows": 0})
    for record in order_records:
        product = clean(record.get("Product Name")) or "未知商品"
        sales = to_decimal(record.get("SKU Subtotal After Discount")) or to_decimal(record.get("Order Amount"))
        stats[product]["units"] = stats[product]["units"] + to_decimal(record.get("Quantity"))  # type: ignore[operator]
        stats[product]["sales"] = stats[product]["sales"] + sales  # type: ignore[operator]
        stats[product]["rows"] = int(stats[product]["rows"]) + 1
    return [[name, int(v["units"]), as_float(v["sales"]), int(v["rows"])] for name, v in sorted(stats.items(), key=lambda x: x[1]["sales"], reverse=True)[:limit]]


def live_ranking(live_records: List[Dict[str, Any]], limit: int = 15) -> List[List[Any]]:
    rows = []
    for record in live_records:
        gmv = to_decimal(record.get("归因 GMV"))
        orders = to_decimal(record.get("归因订单数"))
        views = to_decimal(record.get("观看次数"))
        clicks = to_decimal(record.get("商品点击量"))
        rows.append([record.get("直播日期"), record.get("直播间信息"), record.get("直播开始时间"), as_float(gmv), int(orders), int(views), int(clicks), as_float(div_decimal(orders, views)), as_float(div_decimal(orders, clicks))])
    return sorted(rows, key=lambda row: row[3], reverse=True)[:limit]


def package_rows(packages: List[Package]) -> List[List[Any]]:
    def split_package_id(package_id: str) -> Tuple[str, str, str]:
        parts = package_id.split("-", 2)
        if len(parts) == 3 and parts[0].isalpha() and parts[1].isalpha():
            return parts[0], parts[1], parts[2]
        return "", "", package_id

    rows = []
    for pkg in packages:
        company, platform, serial = split_package_id(pkg.package_id)
        rows.append(
            [
                pkg.package_id,
                company,
                platform,
                serial,
                pkg.package_date.isoformat(),
                pkg.package_rule,
                pkg.tracking_id,
                len(pkg.order_ids) if pkg.order_ids else "",
                as_float(pkg.amount) if pkg.amount else "",
                pkg.customer_name,
                pkg.recipient,
                pkg.zipcode,
                pkg.address,
                "\n".join(pkg.order_ids),
            ]
        )
    return rows


def order_status_rows(orders: List[Order]) -> List[List[Any]]:
    status = Counter(order.status for order in orders)
    substatus = Counter(order.substatus for order in orders)
    return [["Order Status", k, v] for k, v in status.most_common()] + [["Order Substatus", k, v] for k, v in substatus.most_common()]


def fill(cell, color: str) -> None:
    cell.fill = PatternFill("solid", fgColor=color)


def set_default_font(cell) -> None:
    font = copy(cell.font)
    font.name = "Microsoft YaHei"
    cell.font = font
    alignment = copy(cell.alignment)
    alignment.vertical = "center"
    cell.alignment = alignment


def style_title(ws, title: str, subtitle: str, width_cols: int = 14) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=width_cols)
    ws["A1"] = title
    ws["A1"].font = Font(name="Microsoft YaHei", size=20, bold=True, color=PALETTE["yellow"])
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    fill(ws["A1"], PALETTE["blue"])
    ws.row_dimensions[1].height = 30
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=width_cols)
    ws["A2"] = subtitle
    ws["A2"].font = Font(name="Microsoft YaHei", size=10, color=PALETTE["cream"])
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    fill(ws["A2"], PALETTE["red"])
    ws.row_dimensions[2].height = 22


def apply_header_style(ws, row: int, start_col: int, end_col: int) -> None:
    border = Border(bottom=Side(style="thin", color=PALETTE["blue"]))
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row, col)
        fill(cell, PALETTE["yellow"])
        cell.font = Font(name="Microsoft YaHei", bold=True, color=PALETTE["navy"])
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def write_table(ws, start_row: int, start_col: int, headers: Sequence[str], rows: Sequence[Sequence[Any]]) -> int:
    for i, header in enumerate(headers):
        ws.cell(start_row, start_col + i, header)
    apply_header_style(ws, start_row, start_col, start_col + len(headers) - 1)
    for r, row_values in enumerate(rows, start=1):
        for c, value in enumerate(row_values):
            cell = ws.cell(start_row + r, start_col + c, value)
            if r % 2 == 0:
                fill(cell, PALETTE["band"])
    return start_row + len(rows)


def format_currency(ws, range_name: str) -> None:
    for row in ws[range_name]:
        for cell in row:
            cell.number_format = CURRENCY_FORMAT


def format_percent(ws, range_name: str) -> None:
    for row in ws[range_name]:
        for cell in row:
            cell.number_format = PERCENT_FORMAT


def apply_global_number_format(cell) -> None:
    if isinstance(cell.value, (int, float)) and cell.number_format in {"General", "0", "#,##0"}:
        cell.number_format = NUMBER_FORMAT


def set_widths(ws, widths: Dict[str, int]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def style_kpi(ws, row: int, col: int, label: str, value: Any, number_format: str, delta: Optional[Decimal] = None) -> None:
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
    ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 1, end_column=col + 1)
    ws.merge_cells(start_row=row + 2, start_column=col, end_row=row + 2, end_column=col + 1)
    top = ws.cell(row, col, label)
    mid = ws.cell(row + 1, col, value)
    bottom = ws.cell(row + 2, col, "")
    for target in (top, mid, bottom):
        target.alignment = Alignment(horizontal="center", vertical="center")
        target.border = Border(left=Side(style="thin", color=PALETTE["line"]), right=Side(style="thin", color=PALETTE["line"]), top=Side(style="thin", color=PALETTE["line"]), bottom=Side(style="thin", color=PALETTE["line"]))
    fill(top, PALETTE["navy"])
    top.font = Font(name="Microsoft YaHei", bold=True, color=PALETTE["yellow"])
    fill(mid, PALETTE["cream"])
    mid.font = Font(name="Microsoft YaHei", size=15, bold=True, color=PALETTE["black"])
    mid.number_format = number_format
    fill(bottom, PALETTE["sky"])
    bottom.font = Font(name="Microsoft YaHei", size=9, color=PALETTE["gray"])
    if delta is not None:
        direction = "增长" if delta >= 0 else "下降"
        bottom.value = f"较上期{direction} {as_float(abs(delta)):.1%}"
        bottom.font = Font(name="Microsoft YaHei", size=9, bold=True, color=PALETTE["green"] if delta >= 0 else PALETTE["red"])


def delta_rate(current: Decimal, previous: Decimal) -> Optional[Decimal]:
    return None if previous == 0 else (current - previous) / previous


def latest_and_previous(periods: List[PeriodSummary]) -> Tuple[Optional[PeriodSummary], Optional[PeriodSummary]]:
    return (None, None) if not periods else (periods[-1], periods[-2] if len(periods) >= 2 else None)


def add_line_chart(ws, title: str, data_range: str, cat_range: str, anchor: str, y_format: str = NUMBER_FORMAT, color: str = PALETTE["blue"]) -> None:
    chart = LineChart()
    chart.title = title
    chart.height = 7
    chart.width = 15
    chart.style = 13
    chart.y_axis.numFmt = y_format
    chart.y_axis.majorGridlines = None
    chart.legend = None
    chart.add_data(Reference(ws, range_string=f"'{ws.title}'!{data_range}"), titles_from_data=True)
    chart.set_categories(Reference(ws, range_string=f"'{ws.title}'!{cat_range}"))
    if chart.series:
        chart.series[0].graphicalProperties.line.solidFill = color
        chart.series[0].graphicalProperties.line.width = 26000
    ws.add_chart(chart, anchor)


def add_bar_chart(ws, title: str, data_range: str, cat_range: str, anchor: str, color: str = PALETTE["red"]) -> None:
    chart = BarChart()
    chart.title = title
    chart.height = 7
    chart.width = 15
    chart.style = 10
    chart.y_axis.majorGridlines = None
    chart.legend = None
    chart.add_data(Reference(ws, range_string=f"'{ws.title}'!{data_range}"), titles_from_data=True)
    chart.set_categories(Reference(ws, range_string=f"'{ws.title}'!{cat_range}"))
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = False
    if chart.series:
        chart.series[0].graphicalProperties.solidFill = color
    ws.add_chart(chart, anchor)


def create_dashboard(
    wb: Workbook,
    sheet_name: str,
    title: str,
    grain: str,
    periods: List[PeriodSummary],
    products: List[List[Any]],
    lives: List[List[Any]],
    top_customers: List[List[Any]],
    generated_at: datetime,
) -> None:
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = PALETTE["red"] if grain == "day" else PALETTE["blue"] if grain == "week" else PALETTE["yellow"]
    latest, previous = latest_and_previous(periods)
    style_title(ws, title, f"生成时间：{generated_at:%Y-%m-%d %H:%M:%S}    订单：最新全量快照；直播：增量累计", 14)
    if latest is None:
        ws["A4"] = "No data"
        return

    kpis = [
        ("直播 GMV", as_float(latest.live_gmv), CURRENCY_FORMAT, delta_rate(latest.live_gmv, previous.live_gmv) if previous else None),
        ("单场直播GMV", as_float(latest.avg_gmv_per_live), CURRENCY_FORMAT, delta_rate(latest.avg_gmv_per_live, previous.avg_gmv_per_live) if previous else None),
        ("单小时直播GMV", as_float(latest.gmv_per_live_hour), CURRENCY_FORMAT, delta_rate(latest.gmv_per_live_hour, previous.gmv_per_live_hour) if previous else None),
        ("直播场次", latest.live_sessions, NUMBER_FORMAT, delta_rate(Decimal(latest.live_sessions), Decimal(previous.live_sessions)) if previous else None),
        ("订单数", latest.order_count, NUMBER_FORMAT, delta_rate(Decimal(latest.order_count), Decimal(previous.order_count)) if previous else None),
        ("订单/包裹比", as_float(latest.package_order_ratio), NUMBER_FORMAT, delta_rate(latest.package_order_ratio, previous.package_order_ratio) if previous else None),
        ("包裹均价", as_float(latest.avg_package_value), CURRENCY_FORMAT, delta_rate(latest.avg_package_value, previous.avg_package_value) if previous else None),
        ("新客数", latest.new_customers, NUMBER_FORMAT, delta_rate(Decimal(latest.new_customers), Decimal(previous.new_customers)) if previous else None),
        ("老客数", latest.returning_customers, NUMBER_FORMAT, delta_rate(Decimal(latest.returning_customers), Decimal(previous.returning_customers)) if previous else None),
        ("复购客占比", as_float(latest.returning_customer_rate), PERCENT_FORMAT, delta_rate(latest.returning_customer_rate, previous.returning_customer_rate) if previous else None),
        ("退款率", as_float(latest.refund_rate), PERCENT_FORMAT, delta_rate(latest.refund_rate, previous.refund_rate) if previous else None),
        ("待发订单", latest.pending_orders, NUMBER_FORMAT, delta_rate(Decimal(latest.pending_orders), Decimal(previous.pending_orders)) if previous else None),
        ("发货前退款", as_float(latest.pre_ship_refund_amount), CURRENCY_FORMAT, delta_rate(latest.pre_ship_refund_amount, previous.pre_ship_refund_amount) if previous else None),
        ("发货后退款", as_float(latest.post_ship_refund_amount), CURRENCY_FORMAT, delta_rate(latest.post_ship_refund_amount, previous.post_ship_refund_amount) if previous else None),
        ("广告 ROAS", as_float(latest.roas), NUMBER_FORMAT, delta_rate(latest.roas, previous.roas) if previous else None),
    ]
    for idx, item in enumerate(kpis):
        style_kpi(ws, 4 + (idx // 4) * 4, 1 + (idx % 4) * 3, *item)

    rows = [
        [
            item.label,
            as_float(item.live_gmv),
            as_float(item.avg_gmv_per_live),
            as_float(item.gmv_per_live_hour),
            item.order_count,
            item.pending_orders,
            item.packages,
            as_float(item.package_order_ratio),
            as_float(item.avg_package_value),
            item.new_customers,
            item.returning_customers,
            as_float(item.returning_customer_rate),
            as_float(item.pre_ship_refund_amount),
            as_float(item.post_ship_refund_amount),
        ]
        for item in periods[-12:]
    ]
    start = 22
    write_table(ws, start, 1, ["周期", "直播 GMV", "单场直播GMV", "单小时直播GMV", "订单数", "待发订单", "包裹数", "包裹/订单", "包裹均价", "新客", "老客", "复购客占比", "发货前退款", "发货后退款"], rows)
    end = start + len(rows)
    if rows:
        for col in ["B", "C", "D", "I", "M", "N"]:
            format_currency(ws, f"{col}{start+1}:{col}{end}")
        format_percent(ws, f"L{start+1}:L{end}")
        add_line_chart(ws, "GMV 趋势", f"B{start}:B{end}", f"A{start+1}:A{end}", "P4", CURRENCY_FORMAT, PALETTE["blue"])
        add_line_chart(ws, "单小时直播GMV趋势", f"D{start}:D{end}", f"A{start+1}:A{end}", "P20", CURRENCY_FORMAT, PALETTE["red"])

    side = max(end + 3, 36)
    write_table(ws, side, 1, ["Top 用户", "类型", "订单数", "销售额", "退款额", "首购日期"], [[r[3], r[4], r[5], r[6], r[7], r[8]] for r in top_customers])
    if top_customers:
        format_currency(ws, f"D{side+1}:E{side+len(top_customers)}")
        add_bar_chart(ws, "Top 用户销售额", f"D{side}:D{side+len(top_customers)}", f"A{side+1}:A{side+len(top_customers)}", "N36", PALETTE["orange"])
    write_table(ws, side, 8, ["Top 商品", "件数", "销售额", "SKU 行"], products[:8])
    if products:
        format_currency(ws, f"J{side+1}:J{side+len(products[:8])}")
    set_widths(ws, {get_column_letter(i): 12 for i in range(1, 17)})
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["H"].width = 34


def build_workbook(base_dir: Path, sources: SourceBundle, live_records: List[Dict[str, Any]], order_records: List[Dict[str, Any]]) -> Path:
    generated_at = datetime.now()
    orders = order_level_records(order_records)
    table_packages = read_packages_from_table(sources.selected_package_file, infer_report_year(orders))
    tracking_fallback_packages = build_tracking_fallback_packages(orders, table_packages)
    packages = sorted(table_packages + tracking_fallback_packages, key=lambda item: (item.package_date, item.package_id))
    products = product_ranking(order_records)
    lives = live_ranking(live_records)
    daily = build_period_summaries(live_records, orders, packages, "day")
    weekly = build_period_summaries(live_records, orders, packages, "week")
    monthly = build_period_summaries(live_records, orders, packages, "month")

    wb = Workbook()
    wb.remove(wb.active)
    create_dashboard(wb, "日 Dashboard", "TikTok 日报 Dashboard", "day", daily, products, lives, top_customer_rows_for_period(orders, "day", daily[-1] if daily else None), generated_at)
    create_dashboard(wb, "周 Dashboard", "TikTok 周报 Dashboard", "week", weekly, products, lives, top_customer_rows_for_period(orders, "week", weekly[-1] if weekly else None), generated_at)
    create_dashboard(wb, "月 Dashboard", "TikTok 月报 Dashboard", "month", monthly, products, lives, top_customer_rows_for_period(orders, "month", monthly[-1] if monthly else None), generated_at)

    period_headers = ["周期Key", "周期", "开始日期", "结束日期", "直播场次", "单场直播GMV", "单小时直播GMV", "直播GMV", "直播订单", "观看次数", "观看转化", "广告ROAS", "订单数", "订单金额", "退款金额", "退款率", "发货前退款", "发货前退款订单", "发货后退款", "发货后退款订单", "待发订单", "包裹数", "包裹/订单", "包裹均价", "活跃用户", "新客", "老客", "复购客占比", "老客订单", "老客订单占比", "老客销售额"]
    for name, periods in [("日数据", daily), ("周数据", weekly), ("月数据", monthly)]:
        ws = wb.create_sheet(name)
        write_table(ws, 1, 1, period_headers, period_rows(periods))
        ws.freeze_panes = "A2"
        set_widths(ws, {get_column_letter(i): 14 for i in range(1, len(period_headers) + 1)})
        if periods:
            last = len(periods) + 1
            for col in ["F", "G", "H", "N", "O", "Q", "S", "X", "AE"]:
                format_currency(ws, f"{col}2:{col}{last}")
            for col in ["K", "P", "W", "AB", "AD"]:
                format_percent(ws, f"{col}2:{col}{last}")

    pkg_ws = wb.create_sheet("包裹明细")
    write_table(
        pkg_ws,
        1,
        1,
        ["PackageID", "公司", "平台", "流水号", "日期", "包裹来源", "快递单号", "订单数", "包裹成交额", "用户", "收货人", "邮编", "地址", "订单ID"],
        package_rows(packages),
    )
    pkg_ws.freeze_panes = "A2"
    set_widths(pkg_ws, {"A": 22, "B": 10, "C": 10, "D": 14, "E": 12, "F": 18, "G": 20, "H": 10, "I": 14, "J": 18, "K": 22, "L": 12, "M": 54, "N": 70})
    if packages:
        format_currency(pkg_ws, f"I2:I{len(packages)+1}")
        for row in range(2, len(packages) + 2):
            pkg_ws.cell(row, 14).alignment = Alignment(wrap_text=True, vertical="top")

    cust_ws = wb.create_sheet("用户明细")
    customer_rows = customer_period_stats(orders, "day") + customer_period_stats(orders, "week") + customer_period_stats(orders, "month")
    write_table(cust_ws, 1, 1, ["时间维度", "周期Key", "周期", "用户", "类型", "订单数", "销售额", "退款额", "首购日期"], customer_rows)
    cust_ws.freeze_panes = "A2"
    set_widths(cust_ws, {"A": 10, "B": 14, "C": 16, "D": 24, "E": 10, "F": 10, "G": 14, "H": 14, "I": 12})
    if customer_rows:
        format_currency(cust_ws, f"G2:H{len(customer_rows)+1}")

    live_ws = wb.create_sheet("直播明细")
    live_headers = ["经营日期(上海)", "直播间信息", "原始开始时间(LA)", "上海开始时间", "归因 GMV", "归因订单数", "观看次数", "商品点击量", "源文件"]
    live_rows = [[r.get("直播日期"), r.get("直播间信息"), r.get("直播开始时间_原始"), r.get("直播开始时间_上海"), as_float(to_decimal(r.get("归因 GMV"))), int(to_decimal(r.get("归因订单数"))), int(to_decimal(r.get("观看次数"))), int(to_decimal(r.get("商品点击量"))), r.get("Source File")] for r in live_records]
    write_table(live_ws, 1, 1, live_headers, live_rows)
    live_ws.freeze_panes = "A2"
    set_widths(live_ws, {get_column_letter(i): 16 for i in range(1, len(live_headers) + 1)})
    if live_rows:
        format_currency(live_ws, f"D2:D{len(live_rows)+1}")

    order_ws = wb.create_sheet("订单明细")
    order_rows = [[o.order_id, o.order_date.isoformat() if o.order_date else "", o.created_time_raw, o.created_time_report, o.status, o.substatus, as_float(o.amount), as_float(o.refund), "发货后退款" if o.refund > 0 and o.is_shipped_before_refund else "发货前退款" if o.refund > 0 else "", int(o.quantity), o.customer_name, o.recipient, o.zipcode, o.address1, o.source_file] for o in orders]
    write_table(order_ws, 1, 1, ["订单ID", "经营日期(上海)", "原始创建时间(LA)", "上海创建时间", "订单状态", "订单子状态", "订单金额", "退款金额", "退款类型", "件数", "用户", "收货人", "邮编", "地址", "源文件"], order_rows)
    order_ws.freeze_panes = "A2"
    set_widths(order_ws, {"A": 22, "B": 14, "C": 20, "D": 20, "E": 16, "F": 20, "G": 12, "H": 12, "I": 14, "J": 8, "K": 20, "L": 22, "M": 12, "N": 34, "O": 30})
    if order_rows:
        format_currency(order_ws, f"G2:H{len(order_rows)+1}")

    rank_ws = wb.create_sheet("排行榜")
    write_table(rank_ws, 1, 1, ["商品", "件数", "销售额", "SKU行"], products)
    write_table(rank_ws, 1, 7, ["日期", "直播间", "开始时间", "GMV", "订单", "观看", "点击", "观看转化", "点击转化"], lives)
    status_start = max(len(products) + 4, 26)
    write_table(rank_ws, status_start, 1, ["类型", "状态", "订单数"], order_status_rows(orders))
    set_widths(rank_ws, {"A": 52, "B": 12, "C": 12, "D": 10, "G": 12, "H": 18, "I": 18, "J": 12, "K": 10, "L": 10, "M": 10, "N": 12, "O": 12})
    if products:
        format_currency(rank_ws, f"C2:C{len(products)+1}")
        add_bar_chart(rank_ws, "Top 商品销售额", f"C1:C{min(len(products), 10)+1}", f"A2:A{min(len(products), 10)+1}", "Q2", PALETTE["orange"])
    if lives:
        for col in ["J", "K", "L", "M"]:
            for row in range(2, len(lives) + 1):
                rank_ws[f"{col}{row}"].number_format = NUMBER_FORMAT
        for col in ["N", "O"]:
            for row in range(2, len(lives) + 1):
                rank_ws[f"{col}{row}"].number_format = PERCENT_FORMAT
        add_bar_chart(rank_ws, "Top 直播 GMV", f"J1:J{min(len(lives), 10)+1}", f"H2:H{min(len(lives), 10)+1}", "Q18", PALETTE["blue"])

    notes_ws = wb.create_sheet("数据说明")
    notes = [
        ["项目", "内容"],
        ["订单读取口径", "订单文件是全量快照，仅读取源数据文件夹中修改时间最新的一份订单文件。"],
        ["当前订单文件", sources.selected_order_file.name if sources.selected_order_file else "未识别"],
        ["直播读取口径", "直播文件是增量记录，读取所有直播文件，先过滤直播时长低于20分钟的脏数据，再按直播间+开始时间去重。"],
        ["经营日期口径", "TikTok 导出的订单/直播时间按洛杉矶时间理解，并统一转换为上海日期用于日/周/月汇总；包裹表 sheet 日期视为上海日期。"],
        ["Dashboard包裹口径", "日/周/月 Dashboard 的包裹数与订单/包裹比统一基于 All order 订单快照推算：排除取消订单后，同一周期内同收货信息归为一个包裹，单包超过20单时按20单拆分。"],
        ["Dashboard包裹日期口径", "Dashboard 推算包裹按订单创建日期进入日/周/月周期。订单状态会随导出时间变化，因此该口径优先保证可执行和可复现。"],
        ["包裹明细口径", "包裹明细仍保留发货表 PackageID，并在发货表滞后时用 order list 已回写的 Tracking ID 补充临时包裹，作为物流追踪参考，不再驱动 Dashboard 的订单/包裹比。"],
        ["PackageID 规则", "命名规则为“公司-平台-流水号”，例如 VLG-TK-06121。目前公司为 VLG，平台为 TK/TikTok，后续可扩展其他公司和平台。"],
        ["复购口径", "优先用 Buyer Username 识别用户；若缺失则用昵称或收货人+电话。周期内首购为新客，首购早于周期开始为老客。"],
        ["退款口径", "有退款金额且已存在 Shipped Time 或 Delivered Time 记为发货后退款；否则记为发货前退款。"],
        ["直播文件数", len(sources.live_files)],
        ["订单文件数", len(sources.order_files)],
        ["包裹文件数", len(sources.package_files)],
        ["当前包裹文件", sources.selected_package_file.name if sources.selected_package_file else "未识别"],
        ["直播记录数", len(live_records)],
        ["订单唯一数", len(orders)],
        ["发货表包裹数", len(table_packages)],
        ["订单快递单号补充包裹数", len(tracking_fallback_packages)],
        ["包裹数", len(packages)],
        ["脚本版本", SCRIPT_VERSION],
    ]
    write_table(notes_ws, 1, 1, notes[0], notes[1:])
    set_widths(notes_ws, {"A": 18, "B": 110})
    for row in range(2, len(notes) + 1):
        notes_ws.cell(row, 2).alignment = Alignment(wrap_text=True, vertical="top")

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        for row in ws.iter_rows():
            for cell in row:
                set_default_font(cell)
                apply_global_number_format(cell)

    output = base_dir / f"{OUTPUT_PREFIX}_{generated_at.strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(output)
    return output


def run() -> Path:
    script_dir = Path(__file__).resolve().parent
    base_dir = default_data_dir(script_dir)
    base_dir.mkdir(parents=True, exist_ok=True)
    sources = find_sources(base_dir)
    if not sources.live_files and not sources.selected_order_file:
        raise FileNotFoundError(f"源数据文件夹里没有识别到直播或订单文件：{base_dir / SOURCE_DIR_NAME}")
    live_records = read_live_records(sources.live_files)
    order_records = read_order_records(sources.selected_order_file)
    return build_workbook(base_dir, sources, live_records, order_records)


def main() -> int:
    try:
        output = run()
        print("汇报生成完成")
        print(f"输出文件：{output}")
        return 0
    except Exception as exc:
        print(f"汇报生成失败：{exc}")
        return 1


if __name__ == "__main__":
    sys.exit(main())
