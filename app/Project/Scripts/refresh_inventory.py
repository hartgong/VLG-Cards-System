#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
VLG-TCG System V3 - Inventory Refresh, in-place version

This script refreshes the same workbook you pass in:
- Purchase_Inbound -> Inventory_Ledger Purchase In
- Order_Line -> Inventory_Ledger Sales Out
- Inventory_Ledger -> Current_Inventory

Usage:
  python Scripts/refresh_inventory.py Data/card_data_warehouse_V3.xlsx
  python Scripts/refresh_inventory.py Data/card_data_warehouse_V3.xlsx --backup

Important:
- Close the Excel workbook before running this script.
- If no workbook path is provided, the script refreshes ./Data/card_data_warehouse_V3.xlsx.
"""

from __future__ import annotations

import argparse
import shutil
from pathlib import Path
from datetime import datetime
from collections import defaultdict
from typing import Any, Dict, List, Tuple, Optional

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError as e:
    raise SystemExit("缺少 openpyxl，请先运行：pip install openpyxl") from e

PURCHASE_SHEET = "Purchase_Inbound"
LEDGER_SHEET = "Inventory_Ledger"
CURRENT_SHEET = "Current_Inventory"
SKU_SHEET = "SKU_Master"
ORDER_LINE_SHEET = "Order_Line"
ORDER_MASTER_SHEET = "Order_Master"
EXCEPTION_SHEET = "Exception_Log"

LEDGER_HEADERS = [
    "Inventory_Txn_ID", "Txn_Date", "SKU_ID", "SKU_Name", "Txn_Type",
    "Qty_In", "Qty_Out", "Qty_Net", "Unit_Cost_USD", "Amount_USD",
    "Related_Doc_ID", "Related_Order_ID", "Platform", "Company_Code",
    "Host_Name", "Source", "Remark"
]

CURRENT_HEADERS = [
    "SKU_ID", "SKU_Name", "Min_Sales_Unit_Spec", "Standard_Price_USD", "Weight_KG", "Weight_LB",
    "Purchase_In_Qty", "Sales_Out_Qty", "Adjustment_Qty", "Current_Qty",
    "Avg_Unit_Cost_USD", "Inventory_Value_USD", "Inventory_Status", "Note"
]

EXCEPTION_HEADERS = [
    "Exception_ID", "Import Batch ID", "Module", "Severity", "Object_Type",
    "Object_ID", "Exception_Type", "Exception_Detail", "Suggested_Action"
]

# Header aliases make the workbook safer if you later rename a few columns.
ALIASES = {
    "SKU_ID": ["SKU_ID", "SKU", "SKU ID", "商品SKU", "商品编码"],
    "SKU_Name": ["SKU_Name", "Product_Name", "商品名称", "SKU名称"],
    "Product_Name": ["Product_Name", "SKU_Name", "商品名称", "SKU名称"],
    "Purchase_ID": ["Purchase_ID", "采购ID", "采购单号"],
    "Purchase_Date": ["Purchase_Date", "采购日期", "Date"],
    "Purchase_Qty": ["Purchase_Qty", "采购数量", "数量", "Qty", "Quantity"],
    "Unit_Cost_USD": ["Unit_Cost_USD", "采购单价USD", "Unit Cost USD", "单价USD", "成本USD"],
    "Total_Cost_USD": ["Total_Cost_USD", "采购总金额USD", "Total Cost USD", "总金额USD"],
    "Status": ["Status", "状态"],
    "Standard_Order_ID": ["Standard_Order_ID", "标准订单ID"],
    "Order ID": ["Order ID", "Order_ID", "订单ID"],
    "Quantity": ["Quantity", "数量", "Qty"],
}


def header_map(ws) -> Dict[str, int]:
    return {
        str(ws.cell(1, c).value).strip(): c
        for c in range(1, ws.max_column + 1)
        if ws.cell(1, c).value is not None and str(ws.cell(1, c).value).strip() != ""
    }


def resolve_header(hmap: Dict[str, int], header: str) -> Optional[int]:
    if header in hmap:
        return hmap[header]
    for alias in ALIASES.get(header, []):
        if alias in hmap:
            return hmap[alias]
    return None


def get_cell(ws, hmap: Dict[str, int], row: int, header: str, default: Any = None) -> Any:
    col = resolve_header(hmap, header)
    if not col:
        return default
    value = ws.cell(row, col).value
    return default if value is None else value


def safe_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def safe_float(value: Any, default: float = 0.0) -> float:
    if value is None or value == "":
        return default
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip().replace(",", "").replace("$", "")
    if s in ("", "-", "—"):
        return default
    try:
        return float(s)
    except ValueError:
        return default


def clear_data_rows(ws):
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)


def ensure_headers(ws, headers: List[str]):
    for c, h in enumerate(headers, start=1):
        ws.cell(1, c).value = h


def load_sku_master(wb) -> Dict[str, Dict[str, Any]]:
    if SKU_SHEET not in wb.sheetnames:
        return {}
    ws = wb[SKU_SHEET]
    hm = header_map(ws)
    sku_data = {}
    for r in range(2, ws.max_row + 1):
        sku = safe_str(get_cell(ws, hm, r, "SKU_ID"))
        if not sku:
            continue
        sku_data[sku] = {
            "SKU_Name": get_cell(ws, hm, r, "SKU_Name", ""),
            "Min_Sales_Unit_Spec": get_cell(ws, hm, r, "Min_Sales_Unit_Spec", ""),
            "Standard_Price_USD": get_cell(ws, hm, r, "Standard_Price_USD", ""),
            "Weight_KG": get_cell(ws, hm, r, "Weight_KG", ""),
            "Weight_LB": get_cell(ws, hm, r, "Weight_LB", ""),
        }
    return sku_data


def load_order_dates(wb) -> Dict[str, Any]:
    if ORDER_MASTER_SHEET not in wb.sheetnames:
        return {}
    ws = wb[ORDER_MASTER_SHEET]
    hm = header_map(ws)
    result = {}
    for r in range(2, ws.max_row + 1):
        standard_order_id = safe_str(get_cell(ws, hm, r, "Standard_Order_ID"))
        order_id = safe_str(get_cell(ws, hm, r, "Order ID"))
        date_value = get_cell(ws, hm, r, "Created Time Business TZ") or get_cell(ws, hm, r, "Created Time Raw")
        if standard_order_id:
            result[standard_order_id] = date_value
        if order_id:
            result[order_id] = date_value
    return result


def preserve_manual_ledger_rows(wb) -> List[List[Any]]:
    """Keep only manual ledger rows. Auto Purchase In / Sales Out rows are rebuilt from source sheets every run."""
    if LEDGER_SHEET not in wb.sheetnames:
        return []
    ws = wb[LEDGER_SHEET]
    hm = header_map(ws)
    rows = []
    for r in range(2, ws.max_row + 1):
        txn_type = safe_str(get_cell(ws, hm, r, "Txn_Type"))
        source = safe_str(get_cell(ws, hm, r, "Source"))
        if not txn_type:
            continue
        # Auto-generated source rows are always rebuilt.
        if txn_type in {"Purchase In", "Sales Out"} or source in {PURCHASE_SHEET, ORDER_LINE_SHEET}:
            continue
        rows.append([get_cell(ws, hm, r, h, "") for h in LEDGER_HEADERS])
    return rows


def build_purchase_ledger_rows(wb, sku_master: Dict[str, Dict[str, Any]]) -> Tuple[List[List[Any]], List[List[Any]]]:
    rows, exceptions = [], []
    if PURCHASE_SHEET not in wb.sheetnames:
        exceptions.append(["", "Manual/Purchase", "Inventory", "High", "Workbook", PURCHASE_SHEET, "Missing sheet", "Purchase_Inbound sheet not found", "Create Purchase_Inbound sheet"])
        return rows, exceptions

    ws = wb[PURCHASE_SHEET]
    hm = header_map(ws)
    n = 1
    for r in range(2, ws.max_row + 1):
        sku = safe_str(get_cell(ws, hm, r, "SKU_ID"))
        qty = safe_float(get_cell(ws, hm, r, "Purchase_Qty"))
        # A fully blank row is ignored.
        row_values = [ws.cell(r, c).value for c in range(1, min(ws.max_column, 12) + 1)]
        if not any(v not in (None, "") for v in row_values):
            continue
        if not sku and qty == 0:
            continue

        purchase_id = safe_str(get_cell(ws, hm, r, "Purchase_ID")) or f"PUR-ROW-{r}"
        status = safe_str(get_cell(ws, hm, r, "Status"))
        if status.lower() in {"cancelled", "canceled", "void", "invalid", "取消", "作废"}:
            continue

        if not sku:
            exceptions.append(["", "Manual/Purchase", "Inventory", "High", "Purchase", purchase_id, "Missing SKU_ID", f"Purchase_Inbound row {r} has quantity but no SKU_ID", "Fill SKU_ID"])
            continue
        if qty <= 0:
            exceptions.append(["", "Manual/Purchase", "Inventory", "Medium", "Purchase", purchase_id, "Invalid Purchase_Qty", f"Purchase_Inbound row {r} Purchase_Qty <= 0", "Check quantity"])
            continue

        product_name = get_cell(ws, hm, r, "Product_Name") or sku_master.get(sku, {}).get("SKU_Name", "")
        unit_cost = safe_float(get_cell(ws, hm, r, "Unit_Cost_USD"))
        amount_raw = get_cell(ws, hm, r, "Total_Cost_USD")
        amount = safe_float(amount_raw, qty * unit_cost)
        if amount == 0 and unit_cost != 0:
            amount = qty * unit_cost

        rows.append([
            f"INV-PUR-{n:06d}",
            get_cell(ws, hm, r, "Purchase_Date"),
            sku,
            product_name,
            "Purchase In",
            qty,
            0,
            qty,
            unit_cost,
            amount,
            purchase_id,
            "",
            "",
            "",
            "",
            PURCHASE_SHEET,
            "Auto-refreshed in-place from Purchase_Inbound"
        ])
        if sku not in sku_master:
            exceptions.append(["", "Manual/Purchase", "Inventory", "Medium", "SKU", sku, "SKU not in SKU_Master", f"SKU {sku} found in Purchase_Inbound but not in SKU_Master", "Add SKU to SKU_Master"])
        n += 1
    return rows, exceptions


def build_sales_ledger_rows(wb, sku_master: Dict[str, Dict[str, Any]]) -> Tuple[List[List[Any]], List[List[Any]]]:
    rows, exceptions = [], []
    if ORDER_LINE_SHEET not in wb.sheetnames:
        return rows, exceptions

    order_dates = load_order_dates(wb)
    ws = wb[ORDER_LINE_SHEET]
    hm = header_map(ws)
    n = 1
    for r in range(2, ws.max_row + 1):
        sku = safe_str(get_cell(ws, hm, r, "SKU_ID"))
        qty = safe_float(get_cell(ws, hm, r, "Quantity"))
        if not sku and qty == 0:
            continue
        if not sku:
            exceptions.append(["", "Order_Line", "Inventory", "High", "Order_Line", f"row {r}", "Missing SKU_ID", "Order_Line has quantity but no SKU_ID", "Check Seller Note parser"])
            continue
        if qty <= 0:
            exceptions.append(["", "Order_Line", "Inventory", "Medium", "Order_Line", f"row {r}", "Invalid Sales Quantity", "Order_Line Quantity <= 0", "Check Quantity"])
            continue

        standard_order_id = safe_str(get_cell(ws, hm, r, "Standard_Order_ID"))
        order_id = safe_str(get_cell(ws, hm, r, "Order ID"))

        # Critical inventory guardrail:
        # Sales rows for SKU_ID values that do not exist in SKU_Master are invalid.
        # They must NOT be written into Inventory_Ledger, otherwise the system creates
        # ghost SKU inventory records and may deduct stock from non-existent products.
        if sku not in sku_master:
            object_id = standard_order_id or order_id or f"row {r}"
            exceptions.append([
                "", "Order_Line", "Inventory", "High", "Order_Line", object_id,
                "Unknown SKU blocked from Inventory_Ledger",
                f"SKU {sku} found in Order_Line row {r}, but it does not exist in SKU_Master. Sales Out row was NOT added to Inventory_Ledger.",
                "Fix SKU_ID in Order_Line or add the SKU to SKU_Master, then run refresh_inventory.py again"
            ])
            continue

        sku_name = get_cell(ws, hm, r, "SKU_Name") or sku_master.get(sku, {}).get("SKU_Name", "")
        txn_date = order_dates.get(standard_order_id) or order_dates.get(order_id) or ""
        unit_cost = safe_float(get_cell(ws, hm, r, "Unit_Cost_USD"), 0)
        amount = safe_float(get_cell(ws, hm, r, "Total_Product_Cost_USD"), qty * unit_cost if unit_cost else 0)

        rows.append([
            f"INV-SALES-{n:06d}",
            txn_date,
            sku,
            sku_name,
            "Sales Out",
            0,
            qty,
            -qty,
            unit_cost if unit_cost else "",
            amount if amount else "",
            standard_order_id,
            order_id,
            get_cell(ws, hm, r, "Platform", ""),
            get_cell(ws, hm, r, "Company_Code", ""),
            get_cell(ws, hm, r, "Host_Name", ""),
            ORDER_LINE_SHEET,
            "Auto-refreshed in-place from Order_Line"
        ])
        n += 1
    return rows, exceptions


def write_rows(ws, rows: List[List[Any]]):
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(r_idx, c_idx).value = value


def rebuild_current_inventory(wb, sku_master: Dict[str, Dict[str, Any]], ledger_rows: List[List[Any]]):
    ws = wb[CURRENT_SHEET] if CURRENT_SHEET in wb.sheetnames else wb.create_sheet(CURRENT_SHEET)
    ensure_headers(ws, CURRENT_HEADERS)
    clear_data_rows(ws)

    agg = defaultdict(lambda: {"purchase": 0.0, "sales": 0.0, "adjustment": 0.0, "purchase_amount": 0.0})
    sku_names_from_ledger = {}
    all_skus = set(sku_master.keys())
    idx = {h: i for i, h in enumerate(LEDGER_HEADERS)}

    for row in ledger_rows:
        sku = safe_str(row[idx["SKU_ID"]])
        if not sku:
            continue
        all_skus.add(sku)
        if row[idx["SKU_Name"]]:
            sku_names_from_ledger[sku] = row[idx["SKU_Name"]]
        txn_type = safe_str(row[idx["Txn_Type"]])
        qty_in = safe_float(row[idx["Qty_In"]])
        qty_out = safe_float(row[idx["Qty_Out"]])
        qty_net = safe_float(row[idx["Qty_Net"]])
        amount = safe_float(row[idx["Amount_USD"]])

        if txn_type == "Purchase In":
            agg[sku]["purchase"] += qty_in
            agg[sku]["purchase_amount"] += amount
        elif txn_type == "Sales Out":
            agg[sku]["sales"] += qty_out
        else:
            agg[sku]["adjustment"] += qty_net

    rows = []
    for sku in sorted(all_skus):
        info = sku_master.get(sku, {})
        purchase = agg[sku]["purchase"]
        sales = agg[sku]["sales"]
        adjustment = agg[sku]["adjustment"]
        current = purchase - sales + adjustment
        avg_cost = agg[sku]["purchase_amount"] / purchase if purchase else ""
        inv_value = current * avg_cost if avg_cost != "" else ""
        if current < 0:
            status = "Negative Stock"
        elif current == 0:
            status = "Out of Stock"
        elif current <= 5:
            status = "Low Stock"
        else:
            status = "Normal"
        rows.append([
            sku,
            info.get("SKU_Name") or sku_names_from_ledger.get(sku, ""),
            info.get("Min_Sales_Unit_Spec", ""),
            info.get("Standard_Price_USD", ""),
            info.get("Weight_KG", ""),
            info.get("Weight_LB", ""),
            purchase,
            sales,
            adjustment,
            current,
            avg_cost,
            inv_value,
            status,
            "Auto-refreshed from Inventory_Ledger"
        ])
    write_rows(ws, rows)


def reset_inventory_exceptions(wb):
    if EXCEPTION_SHEET not in wb.sheetnames:
        ws = wb.create_sheet(EXCEPTION_SHEET)
        ensure_headers(ws, EXCEPTION_HEADERS)
        return
    ws = wb[EXCEPTION_SHEET]
    ensure_headers(ws, EXCEPTION_HEADERS)
    hm = header_map(ws)
    module_col = resolve_header(hm, "Module")
    if not module_col:
        return
    for r in range(ws.max_row, 1, -1):
        if safe_str(ws.cell(r, module_col).value) == "Inventory":
            ws.delete_rows(r, 1)


def append_exceptions(wb, exceptions: List[List[Any]]):
    if EXCEPTION_SHEET not in wb.sheetnames:
        ws = wb.create_sheet(EXCEPTION_SHEET)
        ensure_headers(ws, EXCEPTION_HEADERS)
    ws = wb[EXCEPTION_SHEET]
    if not exceptions:
        return
    start = ws.max_row + 1
    for offset, row in enumerate(exceptions):
        r_idx = start + offset
        row[0] = f"EXC-INV-{r_idx-1:06d}"
        for c_idx, value in enumerate(row, start=1):
            ws.cell(r_idx, c_idx).value = value


def style_sheet(ws):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(bottom=Side(style="thin", color="D9E2F3"))
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    widths = {
        "A": 20, "B": 18, "C": 18, "D": 30, "E": 15, "F": 12, "G": 12, "H": 12,
        "I": 14, "J": 14, "K": 25, "L": 22, "M": 14, "N": 14, "O": 14, "P": 18, "Q": 40
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def resolve_workbook_path(path_arg: Optional[str]) -> Path:
    if path_arg:
        return Path(path_arg).expanduser().resolve()

    script_path = Path(__file__).resolve()
    repo_root = script_path.parents[3] if len(script_path.parents) >= 4 else Path.cwd()
    default_path = repo_root / "data" / "Project" / "Data" / "card_data_warehouse_V3.xlsx"
    if default_path.exists():
        return default_path.resolve()
    raise SystemExit("未找到默认工作簿。请这样运行：python Scripts/refresh_inventory.py Data/card_data_warehouse_V3.xlsx")


def refresh_inventory_in_place(workbook_path: Path, make_backup: bool = False):
    workbook_path = workbook_path.resolve()
    if not workbook_path.exists():
        raise FileNotFoundError(f"找不到文件：{workbook_path}")
    if workbook_path.suffix.lower() != ".xlsx":
        raise ValueError("请传入 .xlsx 文件")

    print(f"Workbook to refresh: {workbook_path}")
    if make_backup:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = workbook_path.with_name(f"{workbook_path.stem}_backup_{ts}{workbook_path.suffix}")
        shutil.copy2(workbook_path, backup_path)
        print(f"Backup created: {backup_path}")

    wb = load_workbook(workbook_path)
    sku_master = load_sku_master(wb)
    manual_rows = preserve_manual_ledger_rows(wb)
    purchase_rows, purchase_exceptions = build_purchase_ledger_rows(wb, sku_master)
    sales_rows, sales_exceptions = build_sales_ledger_rows(wb, sku_master)
    ledger_rows = purchase_rows + sales_rows + manual_rows

    ws_ledger = wb[LEDGER_SHEET] if LEDGER_SHEET in wb.sheetnames else wb.create_sheet(LEDGER_SHEET)
    ensure_headers(ws_ledger, LEDGER_HEADERS)
    clear_data_rows(ws_ledger)
    write_rows(ws_ledger, ledger_rows)

    rebuild_current_inventory(wb, sku_master, ledger_rows)
    reset_inventory_exceptions(wb)
    append_exceptions(wb, purchase_exceptions + sales_exceptions)

    for sheet_name in [LEDGER_SHEET, CURRENT_SHEET, EXCEPTION_SHEET]:
        if sheet_name in wb.sheetnames:
            style_sheet(wb[sheet_name])

    for ws in [wb[LEDGER_SHEET], wb[CURRENT_SHEET]]:
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                if isinstance(cell.value, float):
                    cell.number_format = "0.00"

    try:
        wb.save(workbook_path)
    except PermissionError as e:
        raise PermissionError("保存失败：请先关闭 Excel 中打开的这个文件，然后重新运行脚本。") from e

    current_ws = wb[CURRENT_SHEET]
    hm = header_map(current_ws)
    total_current = 0.0
    active_rows = 0
    negative = []
    for r in range(2, current_ws.max_row + 1):
        sku = safe_str(get_cell(current_ws, hm, r, "SKU_ID"))
        current = safe_float(get_cell(current_ws, hm, r, "Current_Qty"))
        if sku and current != 0:
            active_rows += 1
        total_current += current
        if current < 0:
            negative.append(sku)

    print("Done. Refreshed in-place.")
    print(f"Purchase In rows rebuilt: {len(purchase_rows)}")
    print(f"Sales Out rows rebuilt: {len(sales_rows)}")
    print(f"Manual ledger rows preserved: {len(manual_rows)}")
    print(f"Inventory_Ledger total rows: {len(ledger_rows)}")
    print(f"Current_Inventory total qty: {total_current:.2f}")
    print(f"SKUs with non-zero inventory: {active_rows}")
    if negative:
        print("Negative stock SKUs:", ", ".join(negative))


def parse_args():
    parser = argparse.ArgumentParser(description="Refresh Inventory_Ledger and Current_Inventory in the same V3 Excel workbook.")
    parser.add_argument("workbook", nargs="?", default=None, help="要原地刷新的 V3 Excel 文件路径，例如 Data/card_data_warehouse_V3.xlsx")
    parser.add_argument("--backup", action="store_true", help="刷新前创建一个备份文件；默认不创建新文件")
    return parser.parse_args()


if __name__ == "__main__":
    args = parse_args()
    path = resolve_workbook_path(args.workbook)
    refresh_inventory_in_place(path, make_backup=args.backup)
