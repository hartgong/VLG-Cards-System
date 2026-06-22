#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Append a small Purchase_Inbound row, then refresh inventory in-place. For testing only."""
from pathlib import Path
from datetime import datetime
import argparse
from openpyxl import load_workbook
from refresh_inventory import refresh_inventory_in_place


def header_map(ws):
    return {str(ws.cell(1, c).value).strip(): c for c in range(1, ws.max_column + 1) if ws.cell(1,c).value is not None}


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", help="V3 Excel workbook path")
    parser.add_argument("--sku", default="TCG00001")
    parser.add_argument("--qty", type=float, default=1)
    parser.add_argument("--unit-cost", type=float, default=0.10)
    args = parser.parse_args()

    path = Path(args.workbook).resolve()
    wb = load_workbook(path)
    ws = wb["Purchase_Inbound"]
    hm = header_map(ws)
    r = ws.max_row + 1
    purchase_id = "PUR-TEST-" + datetime.now().strftime("%Y%m%d%H%M%S")
    values = {
        "Purchase_ID": purchase_id,
        "Purchase_Date": datetime.now().strftime("%Y-%m-%d"),
        "Supplier": "TEST",
        "SKU_ID": args.sku,
        "Product_Name": "TEST PURCHASE",
        "Purchase_Qty": args.qty,
        "Unit_Cost_USD": args.unit_cost,
        "Total_Cost_USD": args.qty * args.unit_cost,
        "Warehouse": "TEST",
        "Status": "Arrived",
        "Remark": "Auto-added by test_add_small_purchase.py",
    }
    for k, v in values.items():
        if k in hm:
            ws.cell(r, hm[k]).value = v
    wb.save(path)
    print(f"Added test purchase row: {purchase_id}, SKU={args.sku}, Qty={args.qty}")
    refresh_inventory_in_place(path, make_backup=False)


if __name__ == "__main__":
    main()
