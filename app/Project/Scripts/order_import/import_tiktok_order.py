from __future__ import annotations

IMPORT_TIKTOK_ORDER_VERSION = "2026-05-27-encoding-fix-v2"

import argparse, csv, sys
from pathlib import Path
from collections import defaultdict
from datetime import datetime
from openpyxl import load_workbook
sys.path.append(str(Path(__file__).resolve().parents[1]))
from common.seller_note_parser import parse_seller_note
from common.time_utils import parse_datetime_to_standard

try:
    import pandas as pd
except ImportError:
    pd = None

ORDER_MASTER_COLUMNS = [
    "Platform","Shop_Name","Order ID","Standard_Order_ID","Order Status",
    "Created Time Raw","Created Time UTC","Created Time Business TZ","Source_Timezone","Parsed_Time_Status",
    "Order Amount_USD","Order Refund Amount_USD","Buyer Username","Country","State","City",
    "Weight_KG","Seller Note Raw","Company_Code","Host_Name","Seller_Note_Parse_Status","Seller_Note_Parse_Note",
    "Import Batch ID","Source File","Data Quality Flag"
]

ORDER_LINE_COLUMNS = [
    "Order_Line_ID","Platform","Shop_Name","Standard_Order_ID","Order ID","Company_Code","Host_Name",
    "SKU_ID","SKU_Name","Quantity","Unit_Price_USD","Line_GMV_USD","Unit_Cost_USD","Total_Product_Cost_USD",
    "Seller_Note_Line","Parse_Source","Parse_Status","Parse_Note","Import Batch ID"
]

INVENTORY_LEDGER_COLUMNS = [
    "Inventory_Txn_ID","Txn_Date","SKU_ID","SKU_Name","Txn_Type","Qty_In","Qty_Out","Qty_Net",
    "Unit_Cost_USD","Amount_USD","Related_Doc_ID","Related_Order_ID","Platform","Company_Code","Host_Name","Source","Remark"
]

EXCEPTION_COLUMNS = ["Exception_ID","Import Batch ID","Module","Severity","Object_Type","Object_ID","Exception_Type","Exception_Detail","Suggested_Action"]

def clean(x):
    return "" if x is None else str(x).replace("\t", "").strip()

def num(x):
    s = clean(x).replace("$", "").replace(",", "")
    if s in ("", "-", "--"):
        return 0.0
    try:
        return float(s)
    except Exception:
        return 0.0

CSV_ENCODINGS_TO_TRY = [
    "utf-8-sig",
    "utf-8",
    "utf-16",
    "gb18030",   # 常见中文 Windows/Excel 导出编码，兼容 GBK/GB2312
    "cp932",     # 日文 Windows/Excel 导出编码
    "shift_jis",
    "big5",
    "latin1",    # 最后兜底，避免脚本因单个特殊字符中断
]


def read_csv(path):
    """读取 TikTok CSV，并自动尝试多种常见编码。

    TikTok/Excel 导出的 CSV 可能不是 UTF-8；如果只按 utf-8-sig 读取，容易出现
    UnicodeDecodeError。这里会依次尝试多种编码，并要求至少能读到 Order ID 表头，
    避免 latin1 兜底时虽然不报错但把表头读乱码。
    """
    path = Path(path)
    decode_errors = []
    last_fieldnames = []

    for encoding in CSV_ENCODINGS_TO_TRY:
        try:
            with open(path, encoding=encoding, newline="", errors="strict") as f:
                reader = csv.DictReader(f)
                rows = list(reader)
                fieldnames = [clean(x) for x in (reader.fieldnames or [])]
        except UnicodeDecodeError as exc:
            decode_errors.append(f"{encoding}: {exc}")
            continue
        except Exception as exc:
            decode_errors.append(f"{encoding}: {type(exc).__name__}: {exc}")
            continue

        last_fieldnames = fieldnames
        if "Order ID" in fieldnames:
            return rows, encoding

    detail = "\n".join(decode_errors[-5:])
    raise ValueError(
        "无法读取 TikTok CSV：没有找到必要表头 'Order ID'。\n"
        f"最后一次读到的表头: {last_fieldnames}\n"
        f"最近的解码错误: {detail}"
    )


def default_warehouse_path() -> Path:
    # import_tiktok_order.py 位于 Scripts/order_import/，项目根目录是 parents[2]
    repo_root = Path(__file__).resolve().parents[4]
    return repo_root / "data" / "Project" / "Data" / "card_data_warehouse_V3.xlsx"

def load_sku_master(warehouse_xlsx: str | Path) -> dict[str, dict]:
    """读取 SKU_Master，并返回 {SKU_ID: row_dict}。

    TikTok 订单导入创建 Order_Line 前必须先校验 SKU 是否存在于 SKU_Master。
    不存在的 SKU 只能进入 Exception_Log，不能进入 Order_Line，也不能生成销售扣库流水。
    """
    path = Path(warehouse_xlsx)
    if not path.exists():
        raise FileNotFoundError(f"SKU warehouse not found: {path}")

    wb = load_workbook(path, read_only=True, data_only=True)
    if "SKU_Master" not in wb.sheetnames:
        raise ValueError(f"SKU_Master sheet not found in warehouse: {path}")

    ws = wb["SKU_Master"]
    rows = ws.iter_rows(values_only=True)
    try:
        headers = [clean(v) for v in next(rows)]
    except StopIteration:
        return {}

    if "SKU_ID" not in headers:
        raise ValueError("SKU_Master must contain SKU_ID column")

    sku_map: dict[str, dict] = {}
    for values in rows:
        row = {headers[i]: values[i] if i < len(values) else None for i in range(len(headers))}
        sku_id = clean(row.get("SKU_ID"))
        if sku_id:
            sku_map[sku_id] = row
    return sku_map

def write_xlsx(output_xlsx, sheets):
    if pd is None:
        raise RuntimeError("Install dependencies first: pip install -r requirements.txt")
    out = Path(output_xlsx)
    if out.suffix.lower() != ".xlsx":
        corrected = out.with_suffix(".xlsx")
        print(f"WARNING: --output 是 Excel 导出路径，已自动改为: {corrected}")
        out = corrected
    out.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        for sheet, (cols, rows) in sheets.items():
            pd.DataFrame(rows, columns=cols).to_excel(writer, index=False, sheet_name=sheet)
    return out

def import_tiktok_order(input_csv: str, output_xlsx: str, platform="TikTok", shop="TikTok_US_MAIN", source_tz="America/Los_Angeles", business_tz="America/Los_Angeles", import_batch_id=None, warehouse_xlsx: str | Path | None = None):
    input_path = Path(input_csv)
    if warehouse_xlsx is None:
        warehouse_xlsx = default_warehouse_path()
    sku_master = load_sku_master(warehouse_xlsx)
    if import_batch_id is None:
        import_batch_id = f"TikTok_Order_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    raw_rows, csv_encoding = read_csv(input_path)
    by_order = defaultdict(list)
    for r in raw_rows:
        oid = clean(r.get("Order ID"))
        if oid:
            by_order[oid].append(r)

    order_master_rows, order_line_rows, inventory_rows, exception_rows = [], [], [], []
    inv_seq = 1
    line_seq = defaultdict(int)

    for order_id, group in by_order.items():
        first = group[0]
        standard_order_id = f"{platform}:{shop}:{order_id}"
        time_info = parse_datetime_to_standard(clean(first.get("Created Time")), source_tz, business_tz)
        seller_note_raw = clean(first.get("Seller Note"))
        parsed = parse_seller_note(seller_note_raw)
        company = parsed["company_code"]
        host = parsed["host_name"]
        data_flag = "" if parsed["status"] == "Parsed" else "Seller Note需检查"

        order_master_rows.append([
            platform, shop, order_id, standard_order_id, clean(first.get("Order Status")),
            clean(first.get("Created Time")), time_info["utc"], time_info["business"], source_tz, time_info["status"],
            num(first.get("Order Amount")), num(first.get("Order Refund Amount")), clean(first.get("Buyer Username")),
            clean(first.get("Country")), clean(first.get("State")), clean(first.get("City")), num(first.get("Weight(kg)")),
            seller_note_raw, company, host, parsed["status"], parsed["note"], import_batch_id, input_path.name, data_flag
        ])

        # v3核心：Seller Note SKU数量就是SKU销售数量，不再换算箱/盒/包。
        if parsed["items"]:
            order_total = num(first.get("Order Amount"))
            total_qty = sum(item["quantity"] for item in parsed["items"]) or 1
            for item in parsed["items"]:
                sku = item["sku_id"]
                qty = item["quantity"]

                sku_info = sku_master.get(sku)
                if sku_info is None:
                    exception_rows.append([
                        f"EXC-{len(exception_rows)+1:06d}", import_batch_id, "order_import", "High", "Order_Line",
                        f"{standard_order_id}:{sku}", "Unknown SKU",
                        f"Seller Note line references SKU '{sku}', but it does not exist in SKU_Master. Raw line: {item['seller_note_line']}",
                        "请先在 SKU_Master 创建该 SKU，或修正 Seller Note 后重新导入；该行不会写入 Order_Line，也不会扣库存"
                    ])
                    continue

                sku_name = clean(sku_info.get("SKU_Name"))
                line_gmv = order_total * qty / total_qty
                unit_price = line_gmv / qty if qty else 0
                line_seq[order_id] += 1
                order_line_id = f"{standard_order_id}:{sku}:{line_seq[order_id]}"
                order_line_rows.append([
                    order_line_id, platform, shop, standard_order_id, order_id, company, host,
                    sku, sku_name, qty, unit_price, line_gmv, "", "",
                    item["seller_note_line"], "Seller Note", "Parsed", "Quantity = SKU quantity; SKU exists in SKU_Master", import_batch_id
                ])
                inventory_rows.append([
                    f"INV-SALES-{inv_seq:06d}", time_info["business"] or clean(first.get("Created Time")), sku, sku_name, "Sales Out",
                    0, qty, -qty, "", "", "", standard_order_id, platform, company, host, "Order_Line",
                    f"Sales out from Seller Note; quantity={qty} SKU units"
                ])
                inv_seq += 1
        else:
            # fallback：未按规范记录Seller Note时，仅生成异常，不自动扣库存，避免错扣。
            exception_rows.append([
                f"EXC-{len(exception_rows)+1:06d}", import_batch_id, "order_import", "High", "Order", order_id,
                "Seller Note Unparsed", seller_note_raw, "请按 #Company / #Host / #SKU Quantity 规则补充Seller Note后重新导入"
            ])

    sheets = {
        "Order_Master": (ORDER_MASTER_COLUMNS, order_master_rows),
        "Order_Line": (ORDER_LINE_COLUMNS, order_line_rows),
        "Inventory_Ledger_Sales": (INVENTORY_LEDGER_COLUMNS, inventory_rows),
        "Exception_Log": (EXCEPTION_COLUMNS, exception_rows),
    }
    saved_path = write_xlsx(output_xlsx, sheets)
    print(f"Saved: {saved_path}")
    print(f"Input CSV encoding: {csv_encoding}")
    print(f"SKU warehouse: {warehouse_xlsx}")
    print(f"Orders: {len(order_master_rows)}, order lines: {len(order_line_rows)}, inventory sales rows: {len(inventory_rows)}, exceptions: {len(exception_rows)}")

if __name__ == "__main__":
    print(f"import_tiktok_order.py version: {IMPORT_TIKTOK_ORDER_VERSION}")
    p = argparse.ArgumentParser()
    p.add_argument("--input", required=True)
    p.add_argument("--output", required=True)
    p.add_argument("--platform", default="TikTok")
    p.add_argument("--shop", default="TikTok_US_MAIN")
    p.add_argument("--source-tz", default="America/Los_Angeles")
    p.add_argument("--business-tz", default="America/Los_Angeles")
    p.add_argument("--warehouse", default=str(default_warehouse_path()), help="Path to card_data_warehouse_V3.xlsx for SKU_Master validation")
    args = p.parse_args()
    import_tiktok_order(args.input, args.output, args.platform, args.shop, args.source_tz, args.business_tz, warehouse_xlsx=args.warehouse)
