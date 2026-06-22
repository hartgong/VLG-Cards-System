# -*- coding: utf-8 -*-
"""
TikTok 发货单 -> 申报单生成工具

默认使用方式：
1) 把本脚本、run_tiktok_declaration_generator.bat、申报配置文件.txt、模板文件、发货单放在同一个目录。
2) 发货单固定命名为：tiktok_order_list_发货单.xlsx
3) 双击 run_tiktok_declaration_generator_v20260610_005.bat 即可生成：tiktok_order_list_申报单.xlsx

配置文件：申报配置文件.txt
- 物流支持：JD / SF / DHL；兼容 JF -> SF；配置文件未填写“物流”时默认 JD
- 销售平台默认写入 TIKTOK；如配置文件存在“销售平台”，则优先使用配置文件值。

模板文件默认在同目录查找：JD模板.xlsx / SF模板.xlsx / DHL模板.xlsx
Package ID 写入目标申报单“备注”栏
长 / 宽 / 高 三列保持空白；兼容旧模板的“外箱尺寸”空白列
收件电话统一格式化为：+1 6507690711
日志文件：tiktok_declaration_generator.log（每次运行覆盖旧日志，仅保留当次执行情况）
"""

from __future__ import annotations

import logging
import re
import sys
import traceback
from copy import copy
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

try:
    from openpyxl import load_workbook
except ImportError as exc:
    raise SystemExit(
        "缺少 openpyxl 模块，请先执行：python -m pip install openpyxl"
    ) from exc


INPUT_FILE_NAME = "tiktok_order_list_发货单.xlsx"
OUTPUT_FILE_NAME = "tiktok_order_list_申报单.xlsx"
CONFIG_FILE_NAME = "申报配置文件.txt"
LOG_FILE_NAME = "tiktok_declaration_generator.log"
DEFAULT_PLATFORM = "TIKTOK"
SCRIPT_VERSION = "v20260610_005"

TEMPLATE_BY_LOGISTICS = {
    "JD": "JD模板.xlsx",
    "SF": "SF模板.xlsx",
    "DHL": "DHL模板.xlsx",
}

# 如果实际输入成 JF，按 SF 处理；如需禁用，删除此映射即可。
LOGISTICS_ALIASES = {
    "JF": "SF",
}

CONFIG_REQUIRED_FIELDS = [
    "发货人",
    "发件国家/地区",
    "州/省",
    "市",
    "手机号",
    "地址",
    "公司英文名",
    "客户商品编码",
    "中文申报品名",
    "英文申报品名",
    "原产国/地区",
]

SOURCE_REQUIRED_HEADERS = [
    "订单号",
    "邮编",
    "收件人",
    "收件电话",
    "收件地址",
    "具体门牌号",
]

TARGET_REQUIRED_HEADERS = [
    "客户订单号",
    "销售平台",
    "销售平台订单号",
    "重量",
    "收货人",
    "州/省",  # 目标表内有两个：第 1 个是收件州/省，第 2 个是发件州/省
    "市",    # 目标表内有两个：第 1 个是收件城市，第 2 个是发件城市
    "电话",
    "详细地址",
    "收货地邮编",
    "发货人",
    "发件国家/地区",
    "手机号",
    "地址",
    "公司英文名",
    "客户商品编码",
    "中文申报品名",
    "英文申报品名",
    "数量",
    "FOB单价",
    "原产国/地区",
    "备注",
]


LENGTH_HEADER_CANDIDATES = ["长", "长CM", "长/cm", "长（CM）", "长(CM)", "长度", "Length"]
WIDTH_HEADER_CANDIDATES = ["宽", "宽CM", "宽/cm", "宽（CM）", "宽(CM)", "宽度", "Width"]
HEIGHT_HEADER_CANDIDATES = ["高", "高CM", "高/cm", "高（CM）", "高(CM)", "高度", "Height"]

TEXT_HEADERS = {
    "备注",
    "客户订单号",
    "销售平台订单号",
    "销售平台",
    "收货人",
    "州/省",
    "市",
    "电话",
    "详细地址",
    "收货地邮编",
    "发货人",
    "发件国家/地区",
    "手机号",
    "地址",
    "公司英文名",
    "客户商品编码",
    "中文申报品名",
    "英文申报品名",
    "原产国/地区",
}


def setup_logger(script_dir: Path) -> logging.Logger:
    logger = logging.getLogger("tiktok_declaration_generator")
    logger.setLevel(logging.INFO)
    logger.handlers.clear()

    log_path = script_dir / LOG_FILE_NAME
    file_handler = logging.FileHandler(log_path, mode="w", encoding="utf-8")
    file_handler.setLevel(logging.INFO)
    formatter = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s")
    file_handler.setFormatter(formatter)
    logger.addHandler(file_handler)
    return logger


def default_data_dir(script_dir: Path) -> Path:
    return script_dir.parents[1] / "data" / "申报"


def normalize_header(value: Any) -> str:
    """去掉换行和空格，用于匹配 Excel 表头。"""
    if value is None:
        return ""
    return re.sub(r"\s+", "", str(value)).strip()


def cell_to_str(value: Any) -> str:
    """把单元格值安全转换成字符串，避免订单号/邮编被处理成科学计数法。"""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def first_non_empty_line(value: Any) -> str:
    """多订单号时取第一条非空订单号。"""
    text = cell_to_str(value)
    for part in re.split(r"[\r\n]+", text):
        part = part.strip()
        if part:
            return part
    return ""


def format_receiver_phone(value: Any) -> str:
    """
    收件电话格式化。
    常见源格式：(+1)6507690711 / +16507690711 / 6507690711
    目标格式：+1 6507690711
    """
    text = cell_to_str(value)
    if not text:
        return ""

    digits = re.sub(r"\D", "", text)

    # 美国号码：1 + 10 位号码
    if len(digits) == 11 and digits.startswith("1"):
        return f"+1 {digits[1:]}"

    # 源数据省略国家码时，默认补 +1。
    if len(digits) == 10:
        return f"+1 {digits}"

    # 兼容其他类似 (+86)13800138000 的格式：去掉括号并在区号后加空格。
    match = re.match(r"^\(\+(\d+)\)\s*(.+)$", text)
    if match:
        country_code = match.group(1)
        rest_digits = re.sub(r"\D", "", match.group(2))
        return f"+{country_code} {rest_digits}" if rest_digits else f"+{country_code}"

    # 兜底：只去掉国家码括号，不强行改变无法识别的号码。
    return text.replace("(+", "+").replace(")", "").strip()


def build_header_map(ws, header_row: int = 1) -> Dict[str, List[int]]:
    """返回 {规范化表头: [列号1, 列号2, ...]}，可处理重复表头。"""
    header_map: Dict[str, List[int]] = {}
    for col in range(1, ws.max_column + 1):
        key = normalize_header(ws.cell(header_row, col).value)
        if key:
            header_map.setdefault(key, []).append(col)
    return header_map


def require_header(
    header_map: Dict[str, List[int]],
    header: str,
    occurrence: int = 1,
    sheet_label: str = "表格",
) -> int:
    """按表头名和出现次数返回列号。occurrence 从 1 开始。"""
    key = normalize_header(header)
    cols = header_map.get(key, [])
    if len(cols) < occurrence:
        raise ValueError(
            f"{sheet_label} 缺少表头：{header}"
            + (f"（第 {occurrence} 个）" if occurrence > 1 else "")
        )
    return cols[occurrence - 1]


def require_any_header(
    header_map: Dict[str, List[int]],
    headers: Sequence[str],
    sheet_label: str = "表格",
) -> int:
    """多个候选表头中，任意一个存在即可。"""
    for header in headers:
        key = normalize_header(header)
        cols = header_map.get(key, [])
        if cols:
            return cols[0]
    raise ValueError(f"{sheet_label} 缺少表头：{' / '.join(headers)}")


def resolve_dimension_columns(header_map: Dict[str, List[int]], sheet_label: str = "目标模板") -> Dict[str, int]:
    """
    兼容新版模板的“长/宽/高”三列，也兼容旧版模板的“外箱尺寸”一列。
    返回的列都会保持空白。
    """
    result: Dict[str, int] = {}

    try:
        result["长"] = require_any_header(header_map, LENGTH_HEADER_CANDIDATES, sheet_label)
        result["宽"] = require_any_header(header_map, WIDTH_HEADER_CANDIDATES, sheet_label)
        result["高"] = require_any_header(header_map, HEIGHT_HEADER_CANDIDATES, sheet_label)
        return result
    except ValueError:
        pass

    key = normalize_header("外箱尺寸")
    cols = header_map.get(key, [])
    if cols:
        result["外箱尺寸"] = cols[0]
        return result

    raise ValueError("目标模板缺少表头：长/宽/高 或 外箱尺寸")


def read_config(config_path: Path) -> Dict[str, str]:
    """读取申报配置文件，支持 Tab / 空格 / 中文冒号 / 英文冒号 / 等号 分隔。"""
    encodings = ["utf-8-sig", "utf-8", "gbk", "gb18030"]
    last_error: Optional[Exception] = None
    text: Optional[str] = None

    for enc in encodings:
        try:
            text = config_path.read_text(encoding=enc)
            break
        except UnicodeDecodeError as exc:
            last_error = exc

    if text is None:
        raise ValueError(f"无法读取配置文件编码：{config_path}，错误：{last_error}")

    config: Dict[str, str] = {}
    for line_no, raw_line in enumerate(text.splitlines(), start=1):
        line = raw_line.strip()
        if not line or line.startswith("#"):
            continue

        if "\t" in line:
            key, value = line.split("\t", 1)
        elif "：" in line:
            key, value = line.split("：", 1)
        elif ":" in line:
            key, value = line.split(":", 1)
        elif "=" in line:
            key, value = line.split("=", 1)
        else:
            parts = line.split(None, 1)
            if len(parts) != 2:
                raise ValueError(f"配置文件第 {line_no} 行无法识别分隔符：{raw_line}")
            key, value = parts

        key = normalize_header(key)
        value = value.strip()
        if key:
            config[key] = value

    missing = [field for field in CONFIG_REQUIRED_FIELDS if normalize_header(field) not in config]
    if missing:
        raise ValueError("配置文件缺少字段：" + "、".join(missing))

    return config


def config_value(config: Dict[str, str], key: str, default: str = "") -> str:
    return config.get(normalize_header(key), default)


def parse_logistics(config: Dict[str, str]) -> str:
    logistics_raw = config_value(config, "物流", "JD").strip().upper()
    logistics = LOGISTICS_ALIASES.get(logistics_raw, logistics_raw)
    if logistics not in TEMPLATE_BY_LOGISTICS:
        raise ValueError("配置文件中的物流只支持：JD / SF / DHL（兼容 JF -> SF）")
    return logistics


def find_required_path(file_name: str, base_dirs: Sequence[Path]) -> Path:
    for base in base_dirs:
        path = base / file_name
        if path.exists():
            return path.resolve()
    searched = [str(base / file_name) for base in base_dirs]
    raise FileNotFoundError("找不到文件：" + file_name + "\n已查找：\n" + "\n".join(searched))


def parse_city_state(address_value: Any, zip_code: str = "") -> Tuple[str, str, str]:
    """
    从源表“收件地址”中提取 city/state/country。
    常见格式：
        收件人
        电话
        Munford, Tennessee
        38058
        United States
    返回：(city, state, country)
    """
    lines = [line.strip() for line in cell_to_str(address_value).splitlines() if line.strip()]
    country = ""

    for line in lines:
        low = line.lower()
        if low in {"united states", "usa", "us", "u.s.", "u.s.a."}:
            country = "United States"
            break

    # 优先取邮编上一行作为 city/state
    location_line = ""
    if zip_code:
        zip_norm = zip_code.strip()
        for idx, line in enumerate(lines):
            if line.strip() == zip_norm or line.strip().startswith(zip_norm + "-"):
                if idx > 0:
                    location_line = lines[idx - 1]
                break

    # 找不到时，取含逗号且不是详细门牌的行
    if not location_line:
        for line in lines:
            if "," in line and not re.search(r"\d", line):
                location_line = line
                break

    # 再退一步：取 United States 前一行；如果前一行是邮编，则取再前一行
    if not location_line:
        for idx, line in enumerate(lines):
            if line.lower() in {"united states", "usa", "us", "u.s.", "u.s.a."}:
                if idx >= 1:
                    prev = lines[idx - 1]
                    if zip_code and (prev == zip_code or prev.startswith(zip_code + "-")) and idx >= 2:
                        location_line = lines[idx - 2]
                    else:
                        location_line = prev
                break

    city = ""
    state = ""
    if location_line:
        parts = [p.strip() for p in location_line.split(",") if p.strip()]
        if len(parts) >= 2:
            city = parts[0]
            state = parts[1]
        elif len(parts) == 1:
            city = parts[0]

    return city, state, country or "United States"


def build_detail_address(street: str, city: str, state: str, country: str) -> str:
    """拼接目标表“详细地址”。"""
    parts = [street, city, state, country]
    return ", ".join([p.strip() for p in parts if p and p.strip()])


def copy_cell_style(src_cell, dst_cell) -> None:
    """复制单元格样式。"""
    if src_cell.has_style:
        dst_cell.font = copy(src_cell.font)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.border = copy(src_cell.border)
        dst_cell.alignment = copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format
        dst_cell.protection = copy(src_cell.protection)
    if src_cell.hyperlink:
        dst_cell._hyperlink = copy(src_cell.hyperlink)
    if src_cell.comment:
        dst_cell.comment = copy(src_cell.comment)


def ensure_remark_column(ws) -> None:
    """确保目标模板存在“备注”列。若模板没有该列，则追加到最后一列。"""
    header_map = build_header_map(ws, 1)
    if normalize_header("备注") in header_map:
        return

    new_col = ws.max_column + 1
    ws.cell(1, new_col).value = "备注"

    # 新增备注列尽量沿用前一列样式。
    if new_col > 1:
        copy_cell_style(ws.cell(1, new_col - 1), ws.cell(1, new_col))
        ws.cell(1, new_col).value = "备注"
        if ws.max_row >= 2:
            copy_cell_style(ws.cell(2, new_col - 1), ws.cell(2, new_col))

def clear_and_prepare_target_rows(ws, row_count: int, data_start_row: int = 2) -> None:
    """清空模板示例数据，只保留表头和格式；后续行复制原第 2 行样式。"""
    max_col = ws.max_column
    style_row = data_start_row if ws.max_row >= data_start_row else 1

    prototype_cells = [ws.cell(style_row, col) for col in range(1, max_col + 1)]
    prototype_styles = []
    for cell in prototype_cells:
        prototype_styles.append({
            "font": copy(cell.font),
            "fill": copy(cell.fill),
            "border": copy(cell.border),
            "alignment": copy(cell.alignment),
            "number_format": cell.number_format,
            "protection": copy(cell.protection),
        })
    prototype_height = ws.row_dimensions[style_row].height

    if ws.max_row >= data_start_row:
        ws.delete_rows(data_start_row, ws.max_row - data_start_row + 1)

    # 逐行应用样式，确保没有模板示例值残留。
    for offset in range(row_count):
        row_num = data_start_row + offset
        if prototype_height is not None:
            ws.row_dimensions[row_num].height = prototype_height
        for col in range(1, max_col + 1):
            dst = ws.cell(row_num, col)
            style = prototype_styles[col - 1]
            dst.font = copy(style["font"])
            dst.fill = copy(style["fill"])
            dst.border = copy(style["border"])
            dst.alignment = copy(style["alignment"])
            dst.number_format = style["number_format"]
            dst.protection = copy(style["protection"])


def source_row_is_blank(ws, row_num: int, source_cols: Dict[str, int]) -> bool:
    check_headers = ["订单号", "邮编", "收件人", "收件电话", "收件地址", "具体门牌号"]
    return all(not cell_to_str(ws.cell(row_num, source_cols[h]).value) for h in check_headers)


def generate_declaration(
    input_path: Path,
    platform: str,
    logistics: str,
    config_path: Path,
    template_path: Path,
    output_path: Path,
) -> int:
    """生成申报单，返回写入行数。"""
    config = read_config(config_path)

    src_wb = load_workbook(input_path, data_only=True)
    src_ws = src_wb.active
    src_header_map = build_header_map(src_ws, 1)
    source_cols = {
        header: require_header(src_header_map, header, 1, "源发货单")
        for header in SOURCE_REQUIRED_HEADERS
    }
    source_cols["Package ID"] = require_any_header(src_header_map, ["Package ID", "包裹号"], "源发货单")

    target_wb = load_workbook(template_path)
    target_ws = target_wb.active
    ensure_remark_column(target_ws)
    target_header_map = build_header_map(target_ws, 1)

    # 基础表头检查
    for header in TARGET_REQUIRED_HEADERS:
        require_header(target_header_map, header, 1, "目标模板")
    dimension_cols = resolve_dimension_columns(target_header_map, "目标模板")
    require_header(target_header_map, "备注", 1, "目标模板")
    # 重复表头检查：目标表中第 2 个“州/省”“市”用于发件人信息
    require_header(target_header_map, "州/省", 2, "目标模板")
    require_header(target_header_map, "市", 2, "目标模板")

    target_cols = {
        "备注": require_header(target_header_map, "备注"),
        "客户订单号": require_header(target_header_map, "客户订单号"),
        "销售平台": require_header(target_header_map, "销售平台"),
        "销售平台订单号": require_header(target_header_map, "销售平台订单号"),
        "重量": require_header(target_header_map, "重量"),
        **dimension_cols,
        "收货人": require_header(target_header_map, "收货人"),
        "收件州/省": require_header(target_header_map, "州/省", 1),
        "收件市": require_header(target_header_map, "市", 1),
        "电话": require_header(target_header_map, "电话"),
        "详细地址": require_header(target_header_map, "详细地址"),
        "收货地邮编": require_header(target_header_map, "收货地邮编"),
        "发货人": require_header(target_header_map, "发货人"),
        "发件国家/地区": require_header(target_header_map, "发件国家/地区"),
        "发件州/省": require_header(target_header_map, "州/省", 2),
        "发件市": require_header(target_header_map, "市", 2),
        "手机号": require_header(target_header_map, "手机号"),
        "地址": require_header(target_header_map, "地址"),
        "公司英文名": require_header(target_header_map, "公司英文名"),
        "客户商品编码": require_header(target_header_map, "客户商品编码"),
        "中文申报品名": require_header(target_header_map, "中文申报品名"),
        "英文申报品名": require_header(target_header_map, "英文申报品名"),
        "数量": require_header(target_header_map, "数量"),
        "FOB单价": require_header(target_header_map, "FOB单价"),
        "原产国/地区": require_header(target_header_map, "原产国/地区"),
    }

    source_data_rows: List[int] = []
    for row_num in range(2, src_ws.max_row + 1):
        if not source_row_is_blank(src_ws, row_num, source_cols):
            source_data_rows.append(row_num)

    clear_and_prepare_target_rows(target_ws, len(source_data_rows), data_start_row=2)

    for out_index, src_row in enumerate(source_data_rows, start=2):
        package_id = cell_to_str(src_ws.cell(src_row, source_cols["Package ID"]).value)
        order_id = first_non_empty_line(src_ws.cell(src_row, source_cols["订单号"]).value)
        zip_code = cell_to_str(src_ws.cell(src_row, source_cols["邮编"]).value)
        receiver_name = cell_to_str(src_ws.cell(src_row, source_cols["收件人"]).value)
        receiver_phone = format_receiver_phone(src_ws.cell(src_row, source_cols["收件电话"]).value)
        address_block = src_ws.cell(src_row, source_cols["收件地址"]).value
        street = cell_to_str(src_ws.cell(src_row, source_cols["具体门牌号"]).value)
        city, state, country = parse_city_state(address_block, zip_code)
        detail_address = build_detail_address(street, city, state, country)

        values_by_col = {
            target_cols["备注"]: package_id,
            target_cols["客户订单号"]: order_id,
            target_cols["销售平台"]: platform,
            target_cols["销售平台订单号"]: order_id,
            target_cols["重量"]: None,
            target_cols["收货人"]: receiver_name,
            target_cols["收件州/省"]: state,
            target_cols["收件市"]: city,
            target_cols["电话"]: receiver_phone,
            target_cols["详细地址"]: detail_address,
            target_cols["收货地邮编"]: zip_code,
            target_cols["发货人"]: config_value(config, "发货人"),
            target_cols["发件国家/地区"]: config_value(config, "发件国家/地区"),
            target_cols["发件州/省"]: config_value(config, "州/省"),
            target_cols["发件市"]: config_value(config, "市"),
            target_cols["手机号"]: config_value(config, "手机号"),
            target_cols["地址"]: config_value(config, "地址"),
            target_cols["公司英文名"]: config_value(config, "公司英文名"),
            target_cols["客户商品编码"]: config_value(config, "客户商品编码"),
            target_cols["中文申报品名"]: config_value(config, "中文申报品名"),
            target_cols["英文申报品名"]: config_value(config, "英文申报品名"),
            target_cols["数量"]: None,
            target_cols["FOB单价"]: None,
            target_cols["原产国/地区"]: config_value(config, "原产国/地区"),
        }
        for dim_key in ["长", "宽", "高", "外箱尺寸"]:
            if dim_key in target_cols:
                values_by_col[target_cols[dim_key]] = None

        for col, value in values_by_col.items():
            cell = target_ws.cell(out_index, col)
            cell.value = value
            header_text = normalize_header(target_ws.cell(1, col).value)
            if header_text in TEXT_HEADERS or col in {
                target_cols["备注"],
                target_cols["客户订单号"],
                target_cols["销售平台订单号"],
                target_cols["收货地邮编"],
                target_cols["电话"],
                target_cols["客户商品编码"],
            }:
                cell.number_format = "@"

    target_wb.save(output_path)
    return len(source_data_rows)


def main() -> None:
    script_dir = Path(__file__).resolve().parent
    data_dir = default_data_dir(script_dir)
    data_dir.mkdir(parents=True, exist_ok=True)
    logger = setup_logger(data_dir)

    try:
        base_dirs = [data_dir, Path.cwd().resolve(), script_dir]
        input_path = find_required_path(INPUT_FILE_NAME, base_dirs)
        config_path = find_required_path(CONFIG_FILE_NAME, [script_dir, input_path.parent.resolve(), *base_dirs])
        config = read_config(config_path)
        logistics = parse_logistics(config)
        platform = config_value(config, "销售平台", DEFAULT_PLATFORM).strip() or DEFAULT_PLATFORM
        template_path = find_required_path(TEMPLATE_BY_LOGISTICS[logistics], [script_dir, input_path.parent.resolve(), *base_dirs])
        output_path = data_dir / OUTPUT_FILE_NAME

        if output_path.exists():
            try:
                output_path.unlink()
            except PermissionError as exc:
                raise PermissionError(f"输出文件已打开或无权限覆盖，请先关闭文件：{output_path}") from exc

        logger.info("脚本版本：%s", SCRIPT_VERSION)
        logger.info("开始生成申报单")
        logger.info("源文件：%s", input_path)
        logger.info("模板：%s", template_path)
        logger.info("配置：%s", config_path)
        logger.info("物流：%s", logistics)
        logger.info("销售平台：%s", platform)

        count = generate_declaration(
            input_path=input_path,
            platform=platform,
            logistics=logistics,
            config_path=config_path,
            template_path=template_path,
            output_path=output_path,
        )

        logger.info("申报单生成完成，写入行数：%s，输出文件：%s", count, output_path)
        print(f"脚本版本：{SCRIPT_VERSION}")
        print("申报单生成完成")
        print(f"源文件：{input_path}")
        print(f"模板：{template_path}")
        print(f"配置：{config_path}")
        print(f"物流：{logistics}")
        print(f"销售平台：{platform}")
        print(f"写入包裹/订单行数：{count}")
        print(f"输出文件：{output_path}")

    except Exception as exc:
        error_text = "运行失败：" + str(exc)
        logger.error(error_text)
        logger.error(traceback.format_exc())
        print(error_text)
        print(f"详细错误已写入：{script_dir / LOG_FILE_NAME}")
        sys.exit(1)


if __name__ == "__main__":
    main()
