#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
KQ TikTok settlement generator.

It reads a monthly data folder, pairs each store's All order + income files,
keeps only KQ settlement rows, and writes a Pokemon-style P&L workbook.
"""

from __future__ import annotations

import re
from copy import copy
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Set, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


USD_TO_RMB = Decimal("6.80")
VLG_COMMISSION_RATE = Decimal("0.03")
SETTLEMENT_MONTH = "2026-05"
KQ_KEYWORD = "kq"
FORMAT_TEMPLATE_NAME = "KQ_TikTok_结算单_模板.xlsx"

ORDER_ID_RE = re.compile(r"\d{15,22}")
MONEY = Decimal("0.01")

PALETTE = {
    "blue": "2A75BB",
    "yellow": "FFCB05",
    "red": "E3350D",
    "navy": "1B2A41",
    "cream": "FFF7D6",
    "sky": "D9ECFF",
    "green": "ECFDF5",
    "green_text": "065F46",
    "line": "D7DEE8",
    "white": "FFFFFF",
    "black": "111827",
    "band": "F8FBFF",
}


def clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, Decimal):
        return str(value)
    return str(value).strip()


def norm_header(value: Any) -> str:
    return re.sub(r"\s+", " ", clean(value).replace("\n", " ")).strip().lower()


def dec(value: Any) -> Decimal:
    text = clean(value).replace(",", "")
    if not text:
        return Decimal("0")
    try:
        return Decimal(text)
    except InvalidOperation:
        return Decimal("0")


def money(value: Decimal) -> Decimal:
    return value.quantize(MONEY, rounding=ROUND_HALF_UP)


def as_float(value: Decimal) -> float:
    return float(money(value))


def extract_order_ids(*values: Any) -> Set[str]:
    ids: Set[str] = set()
    for value in values:
        ids.update(ORDER_ID_RE.findall(clean(value)))
    return ids


def normalized_order_id(value: Any) -> str:
    ids = extract_order_ids(value)
    return next(iter(ids), "")


def in_settlement_month(value: Any) -> bool:
    text = clean(value)
    return text.startswith(SETTLEMENT_MONTH.replace("-", "/")) or text.startswith(SETTLEMENT_MONTH)


def header_map(headers: Sequence[Any]) -> Dict[str, int]:
    return {norm_header(header): idx for idx, header in enumerate(headers)}


def col(headers: Dict[str, int], *names: str) -> int:
    for name in names:
        key = norm_header(name)
        if key in headers:
            return headers[key]
    raise KeyError(f"缺少必要列：{names}")


def find_col_contains(headers: Sequence[Any], *needles: str) -> int:
    candidates = [(idx, norm_header(header)) for idx, header in enumerate(headers)]
    for needle in needles:
        target = norm_header(needle)
        for idx, header in candidates:
            if target in header:
                return idx
    raise KeyError(f"缺少必要列：{needles}")


def find_month_dir(root: Path) -> Path:
    if any(p.suffix.lower() == ".xlsx" and not p.name.startswith("~$") for p in root.iterdir() if p.is_file()):
        return root
    dirs = [p for p in root.iterdir() if p.is_dir() and any(x.suffix.lower() == ".xlsx" for x in p.iterdir() if x.is_file())]
    if not dirs:
        raise FileNotFoundError(f"未找到月份数据文件夹：{root}")
    return max(dirs, key=lambda p: p.stat().st_mtime)


def display_store_name(name: str) -> str:
    name = re.sub(r"\s+All order.*$", "", name, flags=re.I)
    name = re.sub(r"\s+income.*$", "", name, flags=re.I)
    name = re.sub(r"\s+2026-?5$", "", name, flags=re.I).strip()
    return name


@dataclass
class StoreFiles:
    store: str
    order_file: Path
    income_file: Path


@dataclass
class StoreTotals:
    store: str
    order_ids: Set[str] = field(default_factory=set)
    total_settlement_usd: Decimal = Decimal("0")
    net_sales_usd: Decimal = Decimal("0")
    shipping_usd: Decimal = Decimal("0")
    promotion_usd_raw: Decimal = Decimal("0")
    referral_usd_raw: Decimal = Decimal("0")

    @property
    def revenue_usd(self) -> Decimal:
        return self.net_sales_usd + self.shipping_usd

    @property
    def revenue_rmb(self) -> Decimal:
        return self.revenue_usd * USD_TO_RMB

    @property
    def commission_rmb(self) -> Decimal:
        return self.revenue_rmb * VLG_COMMISSION_RATE

    @property
    def promotion_rmb(self) -> Decimal:
        return abs(self.promotion_usd_raw) * USD_TO_RMB

    @property
    def referral_rmb(self) -> Decimal:
        return abs(self.referral_usd_raw) * USD_TO_RMB

    @property
    def final_receipt_rmb(self) -> Decimal:
        return self.total_settlement_usd * USD_TO_RMB


@dataclass
class StoreData:
    files: StoreFiles
    totals: StoreTotals
    income_headers: List[Any]
    income_rows: List[List[Any]]
    order_headers: List[Any]
    order_rows: List[List[Any]]
    abnormal_rows: List[List[Any]]
    kq_source_note: str


@dataclass
class PackageCost:
    jdw_no: str
    order_ids: Set[str] = field(default_factory=set)
    tax_rmb: Decimal = Decimal("0")
    linehaul_rmb: Decimal = Decimal("0")
    registration_rmb: Decimal = Decimal("0")
    other_rmb: Decimal = Decimal("0")


def pair_store_files(base_dir: Path) -> List[StoreFiles]:
    order_files = [p for p in base_dir.glob("*.xlsx") if "all order" in p.name.lower() and not p.name.startswith("~$")]
    income_files = [p for p in base_dir.glob("*.xlsx") if "income" in p.name.lower() and not p.name.startswith("~$")]
    stores: List[StoreFiles] = []
    for income in sorted(income_files, key=lambda p: p.name):
        income_key = display_store_name(income.name).lower()
        best: Optional[Path] = None
        best_score = -1
        for order in order_files:
            order_key = display_store_name(order.name).lower()
            score = len(set(income_key.split()) & set(order_key.split()))
            if score > best_score:
                best = order
                best_score = score
        if best is None:
            raise FileNotFoundError(f"未找到 {income.name} 对应的 All order 文件")
        stores.append(StoreFiles(store=display_store_name(best.name), order_file=best, income_file=income))
    return stores


def load_all_order_selection(order_file: Path) -> Tuple[List[Any], List[List[Any]], Set[str], List[List[Any]], bool]:
    # Some TikTok exports have a stale worksheet dimension of A1:A1 even though
    # Excel displays all rows. Normal mode reads the real cells reliably.
    wb = load_workbook(order_file, read_only=False, data_only=True)
    ws = wb.worksheets[0]
    rows = ws.iter_rows(values_only=True)
    headers = list(next(rows))
    if len(headers) <= 1:
        return headers, [], set(), [], True
    h = header_map(headers)
    c_order = col(h, "Order ID")
    c_product = col(h, "Product Name")
    c_note = col(h, "Seller Note")
    c_cancel_return = col(h, "Cancelation/Return Type")
    c_status = col(h, "Order Status")
    c_substatus = col(h, "Order Substatus")
    all_rows: List[List[Any]] = []
    kq_ids: Set[str] = set()
    abnormal_rows: List[List[Any]] = []
    has_data = False
    for row_number, row in enumerate(rows, start=2):
        if row_number == 2:
            continue
        order_id = normalized_order_id(row[c_order])
        if not order_id:
            continue
        has_data = True
        row_values = list(row)
        all_rows.append(row_values)
        product_hit = KQ_KEYWORD in clean(row[c_product]).lower()
        note_hit = KQ_KEYWORD in clean(row[c_note]).lower()
        if product_hit or note_hit:
            kq_ids.add(order_id)
            cancel_return_type = clean(row[c_cancel_return])
            if cancel_return_type:
                abnormal_rows.append([
                    "取消/退款异常",
                    order_id,
                    f"Cancelation/Return Type={cancel_return_type}; "
                    f"Order Status={clean(row[c_status])}; "
                    f"Order Substatus={clean(row[c_substatus])}",
                ])
    return headers, all_rows, kq_ids, abnormal_rows, not has_data


def load_income_for_store(
    income_file: Path,
    selected_order_ids: Set[str],
    include_all_if_no_order_data: bool,
    store: str,
    refund_order_ids: Set[str],
) -> Tuple[List[Any], List[List[Any]], StoreTotals]:
    wb = load_workbook(income_file, read_only=True, data_only=True)
    ws = wb["Order details"] if "Order details" in wb.sheetnames else wb.worksheets[0]
    rows = ws.iter_rows(values_only=True)
    headers = list(next(rows))
    h = header_map(headers)
    c_statement = col(h, "Statement date")
    c_type = col(h, "Type")
    c_order = col(h, "Order/adjustment ID")
    c_total = col(h, "Total settlement amount")
    c_net_sales = col(h, "Net sales")
    c_shipping = col(h, "Shipping")
    c_referral = col(h, "Referral fee")
    c_smart = col(h, "Smart Promotion fee")
    c_smart_period = col(h, "Smart Promotion campaign period fee")

    detail_rows: List[List[Any]] = []
    totals = StoreTotals(store=store)
    for row in rows:
        order_id = normalized_order_id(row[c_order])
        if not order_id:
            continue
        if clean(row[c_type]).lower() != "order":
            continue
        if not in_settlement_month(row[c_statement]):
            continue
        if not include_all_if_no_order_data and order_id not in selected_order_ids:
            continue
        if order_id in refund_order_ids:
            continue
        detail_rows.append(list(row))
        totals.order_ids.add(order_id)
        totals.total_settlement_usd += dec(row[c_total])
        totals.net_sales_usd += dec(row[c_net_sales])
        totals.shipping_usd += dec(row[c_shipping])
        totals.referral_usd_raw += dec(row[c_referral])
        totals.promotion_usd_raw += dec(row[c_smart]) + dec(row[c_smart_period])
    return headers, detail_rows, totals


def filter_order_rows(headers: List[Any], all_rows: List[List[Any]], order_ids: Set[str]) -> List[List[Any]]:
    if not all_rows or len(headers) <= 1:
        return []
    c_order = col(header_map(headers), "Order ID")
    return [row for row in all_rows if normalized_order_id(row[c_order]) in order_ids]


def load_store_data(files: StoreFiles, refund_order_ids: Set[str], refund_reasons: Dict[str, str]) -> StoreData:
    order_headers, all_order_rows, kq_ids, abnormal_rows, no_order_data = load_all_order_selection(files.order_file)
    refund_hits = kq_ids & refund_order_ids
    for order_id in sorted(refund_hits):
        abnormal_rows.append([
            "退款清单异常",
            order_id,
            refund_reasons.get(order_id, "该KQ订单出现在Return/Refund Orders清单中，已从结算汇总剔除"),
        ])
    income_headers, income_rows, totals = load_income_for_store(files.income_file, kq_ids, no_order_data, files.store, refund_order_ids)
    order_rows = filter_order_rows(order_headers, all_order_rows, totals.order_ids)
    note = "All order: Product Name/Seller Note KQ"
    if no_order_data:
        note = "All order无订单行，按该店本月income全量纳入"
    return StoreData(
        files=files,
        totals=totals,
        income_headers=income_headers,
        income_rows=income_rows,
        order_headers=order_headers,
        order_rows=order_rows,
        abnormal_rows=abnormal_rows,
        kq_source_note=note,
    )


def find_jd_file(base_dir: Path) -> Path:
    files = [p for p in base_dir.glob("*.xlsx") if "u4eac" in p.name.encode("unicode_escape").decode() and not p.name.startswith("~$")]
    if not files:
        raise FileNotFoundError("未找到京东物流文件")
    return max(files, key=lambda p: p.stat().st_mtime)


def find_shipment_file(base_dir: Path) -> Path:
    files = [p for p in base_dir.glob("*.xlsx") if "u53d1" in p.name.encode("unicode_escape").decode() and not p.name.startswith("~$")]
    if not files:
        raise FileNotFoundError("未找到发货明细文件")
    return max(files, key=lambda p: p.stat().st_mtime)


def find_refund_files(base_dir: Path) -> List[Path]:
    return [
        p for p in base_dir.glob("*.xlsx")
        if "return" in p.name.lower()
        and "refund" in p.name.lower()
        and not p.name.startswith("~$")
    ]


def load_refund_orders(base_dir: Path) -> Tuple[Set[str], Dict[str, str]]:
    refund_ids: Set[str] = set()
    reasons: Dict[str, str] = {}
    for refund_file in find_refund_files(base_dir):
        wb = load_workbook(refund_file, read_only=False, data_only=True)
        ws = wb.worksheets[0]
        rows = ws.iter_rows(values_only=True)
        headers = list(next(rows))
        h = header_map(headers)
        c_order = col(h, "Order ID")
        c_return_type = h.get(norm_header("Return Type"))
        c_return_status = h.get(norm_header("Return Status"))
        c_return_substatus = h.get(norm_header("Return Sub Status"))
        c_return_reason = h.get(norm_header("Return Reason"))
        for row in rows:
            order_id = normalized_order_id(row[c_order])
            if not order_id:
                continue
            refund_ids.add(order_id)
            detail = [
                f"Return/Refund文件={refund_file.name}",
                f"Return Type={clean(row[c_return_type])}" if c_return_type is not None else "",
                f"Return Status={clean(row[c_return_status])}" if c_return_status is not None else "",
                f"Return Sub Status={clean(row[c_return_substatus])}" if c_return_substatus is not None else "",
                f"Return Reason={clean(row[c_return_reason])}" if c_return_reason is not None else "",
                "已从结算汇总剔除",
            ]
            reasons[order_id] = "; ".join(item for item in detail if item)
    return refund_ids, reasons


def load_shipment_rows(shipment_file: Path) -> Tuple[str, List[Any], List[List[Any]], Set[str]]:
    wb = load_workbook(shipment_file, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows = ws.iter_rows(values_only=True)
    headers = list(next(rows))
    c_order = find_col_contains(headers, "tk")
    detail_rows: List[List[Any]] = []
    order_ids: Set[str] = set()
    for row in rows:
        row_values = list(row)
        detail_rows.append(row_values)
        order_id = normalized_order_id(row[c_order])
        if order_id:
            order_ids.add(order_id)
    return ws.title, headers, detail_rows, order_ids


def load_jd_rows(jd_file: Path, settlement_order_ids: Set[str]) -> Tuple[List[Any], List[List[Any]], Dict[str, PackageCost]]:
    wb = load_workbook(jd_file, read_only=True, data_only=True)
    ws = wb["明细"] if "明细" in wb.sheetnames else wb.worksheets[1]
    rows = ws.iter_rows(values_only=True)
    headers = list(next(rows))
    c_customer_pkg = find_col_contains(headers, "Customer Package No", "商家包裹号")
    c_item = find_col_contains(headers, "Billing Items", "计费项")
    c_amount = find_col_contains(headers, "实际结算金额")
    c_jdw = find_col_contains(headers, "jdwNo")
    try:
        c_sales_order = find_col_contains(headers, "Sales Platform Order Number", "销售平台订单号")
    except KeyError:
        c_sales_order = None

    raw_by_jdw: Dict[str, List[List[Any]]] = defaultdict(list)
    all_packages: Dict[str, PackageCost] = {}
    matched_jdws: Set[str] = set()
    for row in rows:
        jdw_no = clean(row[c_jdw])
        if not jdw_no:
            continue
        row_values = list(row)
        raw_by_jdw[jdw_no].append(row_values)
        ids = extract_order_ids(row[c_customer_pkg], row[c_sales_order] if c_sales_order is not None else "")
        pkg = all_packages.setdefault(jdw_no, PackageCost(jdw_no=jdw_no))
        pkg.order_ids.update(ids)
        amount = dec(row[c_amount])
        item = clean(row[c_item])
        if "关税" in item:
            pkg.tax_rmb += amount
        elif "专线运费" in item:
            pkg.linehaul_rmb += amount
        elif "挂号费" in item:
            pkg.registration_rmb += amount
        else:
            pkg.other_rmb += amount
        if ids & settlement_order_ids:
            matched_jdws.add(jdw_no)

    rows_out: List[List[Any]] = []
    packages: Dict[str, PackageCost] = {}
    for jdw_no in sorted(matched_jdws):
        rows_out.extend(raw_by_jdw[jdw_no])
        packages[jdw_no] = all_packages[jdw_no]
    return headers, rows_out, packages


def allocated_logistics(packages: Dict[str, PackageCost], settlement_order_ids: Set[str]) -> Tuple[Decimal, Decimal, Decimal, Decimal]:
    tax = linehaul = registration = other = Decimal("0")
    for pkg in packages.values():
        relevant = pkg.order_ids & settlement_order_ids
        if not relevant:
            continue
        denominator = Decimal(max(len(pkg.order_ids), 1))
        factor = Decimal(len(relevant)) / denominator
        tax += pkg.tax_rmb * factor
        linehaul += pkg.linehaul_rmb * factor
        registration += pkg.registration_rmb * factor
        other += pkg.other_rmb * factor
    return tax, linehaul, registration, other


def style_header_row(ws, row_idx: int, fill_color: str = PALETTE["blue"]) -> None:
    fill = PatternFill("solid", fgColor=fill_color)
    font = Font(name="Microsoft YaHei", bold=True, color=PALETTE["white"])
    side = Side(style="thin", color=PALETTE["line"])
    border = Border(left=side, right=side, top=side, bottom=side)
    for cell in ws[row_idx]:
        cell.fill = fill
        cell.font = font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_raw_sheet(ws, max_col: int) -> None:
    style_header_row(ws, 1)
    side = Side(style="thin", color=PALETTE["line"])
    border = Border(left=side, right=side, top=side, bottom=side)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Microsoft YaHei", size=10, color=PALETTE["black"])
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = border
            if cell.row % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=PALETTE["band"])
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    for col_idx in range(1, max_col + 1):
        header = clean(ws.cell(1, col_idx).value)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(len(header) + 2, 10), 28)


def write_raw_sheet(wb: Workbook, name: str, headers: Sequence[Any], rows: Sequence[Sequence[Any]], max_width: int = 28) -> None:
    ws = wb.create_sheet(name[:31])
    ws.append(list(headers))
    for row in rows:
        ws.append(list(row))
    style_raw_sheet(ws, max(len(headers), 1))
    for col_idx in range(1, max(len(headers), 1) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = min(ws.column_dimensions[get_column_letter(col_idx)].width or 10, max_width)


def write_summary(wb: Workbook, store_data: List[StoreData], logistics: Tuple[Decimal, Decimal, Decimal, Decimal], base_dir: Path, jd_file: Path, shipment_file: Path) -> None:
    ws = wb.create_sheet("结算汇总")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:E1")
    ws["A1"] = "杭州克勤 TikTok 结算 P&L"
    ws["A1"].font = Font(name="Microsoft YaHei", size=18, bold=True, color=PALETTE["yellow"])
    ws["A1"].fill = PatternFill("solid", fgColor=PALETTE["blue"])
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    meta = [
        ("结算月份", SETTLEMENT_MONTH),
        ("美元兑人民币汇率", as_float(USD_TO_RMB)),
        ("VLG佣金率", float(VLG_COMMISSION_RATE)),
        ("结算对象标志位", KQ_KEYWORD.upper()),
    ]
    for row_idx, (label, value) in enumerate(meta, start=3):
        ws.cell(row_idx, 1, label)
        ws.cell(row_idx, 2, value)
    ws["B5"].number_format = "0.00%"

    stores = [item.totals.store for item in store_data]
    if len(stores) < 2:
        stores.append("店铺2")
    header_row = 9
    headers = ["P&L项目", stores[0], stores[1], "合计RMB", "口径"]
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(header_row, col_idx, header)
    style_header_row(ws, header_row, PALETTE["navy"])

    totals_by_store = {item.totals.store: item.totals for item in store_data}
    def store_values(attr: str) -> List[Decimal]:
        return [getattr(totals_by_store.get(store, StoreTotals(store)), attr) for store in stores[:2]]

    tax, linehaul, registration, other = logistics
    product_cost = Decimal("0")
    revenue_vals = store_values("revenue_rmb")
    commission_vals = store_values("commission_rmb")
    promotion_vals = store_values("promotion_rmb")
    referral_vals = store_values("referral_rmb")
    settlement = sum(revenue_vals, Decimal("0")) - sum(commission_vals, Decimal("0")) - sum(promotion_vals, Decimal("0")) - sum(referral_vals, Decimal("0")) - tax - linehaul - registration - other - product_cost

    rows = [
        ("收入", revenue_vals, sum(revenue_vals, Decimal("0")), "净销售额 = Net sales(P) + Shipping(U)，按店铺区分"),
        ("减：VLG佣金", commission_vals, sum(commission_vals, Decimal("0")), "收入 × 3%，按店铺区分"),
        ("减：推广费", promotion_vals, sum(promotion_vals, Decimal("0")), "Smart Promotion fee(BI) + Smart Promotion campaign period fee(BM)，按店铺区分"),
        ("减：平台扣点", referral_vals, sum(referral_vals, Decimal("0")), "Referral fee(AT)，按店铺区分"),
        ("减：关税", [None, None], tax, "京东费用项=关税，合计列列示"),
        ("减：专线运费", [None, None], linehaul, "京东费用项=专线运费，合计列列示"),
        ("减：挂号费", [None, None], registration, "京东费用项=挂号费，合计列列示"),
        ("减：其他物流", [None, None], other, "京东其他费用项，合计列列示"),
        ("减：商品成本", [None, None], product_cost, ""),
        ("KQ结算金额", [None, None], settlement, "收入 - VLG佣金 - 推广费 - 平台扣点 - 物流/关税 - 商品成本"),
    ]
    for row_idx, (label, vals, total, note) in enumerate(rows, start=header_row + 1):
        ws.cell(row_idx, 1, label)
        for i, value in enumerate(vals[:2], start=2):
            ws.cell(row_idx, i, "" if value is None else as_float(value))
        ws.cell(row_idx, 4, as_float(total))
        ws.cell(row_idx, 5, note)

    side = Side(style="thin", color=PALETTE["line"])
    border = Border(left=side, right=side, top=side, bottom=side)
    for row in ws.iter_rows(min_row=3, max_row=header_row + len(rows), min_col=1, max_col=5):
        for cell in row:
            cell.font = Font(name="Microsoft YaHei", size=10, color=PALETTE["black"])
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = border
    for row_idx in range(header_row + 1, header_row + len(rows) + 1):
        if ws.cell(row_idx, 1).value in {"收入", "KQ结算金额"}:
            for col_idx in range(1, 6):
                ws.cell(row_idx, col_idx).fill = PatternFill("solid", fgColor=PALETTE["green"])
                ws.cell(row_idx, col_idx).font = Font(name="Microsoft YaHei", bold=True, color=PALETTE["green_text"])
        for col_idx in [2, 3, 4]:
            ws.cell(row_idx, col_idx).number_format = "#,##0.00"
    ws["B4"].number_format = "#,##0.00"
    ws["B5"].number_format = "0.00%"
    for col, width in {"A": 22, "B": 30, "C": 18, "D": 16, "E": 68}.items():
        ws.column_dimensions[col].width = width


def apply_format_template(wb: Workbook, base_dir: Path) -> None:
    template = base_dir / FORMAT_TEMPLATE_NAME
    if not template.exists():
        raise FileNotFoundError(f"固定格式模板不存在：{template}")
    src_wb = load_workbook(template, data_only=False)
    for ws in wb.worksheets:
        if ws.title not in src_wb.sheetnames:
            continue
        src = src_wb[ws.title]
        ws.sheet_view.showGridLines = src.sheet_view.showGridLines
        ws.freeze_panes = src.freeze_panes
        for key, dim in src.column_dimensions.items():
            ws.column_dimensions[key].width = dim.width
            ws.column_dimensions[key].hidden = dim.hidden
        for key, dim in src.row_dimensions.items():
            ws.row_dimensions[key].height = dim.height
            ws.row_dimensions[key].hidden = dim.hidden
        max_row = max(ws.max_row, src.max_row)
        max_col = max(ws.max_column, src.max_column)
        for row in range(1, max_row + 1):
            for col_idx in range(1, max_col + 1):
                src_cell = src.cell(row, col_idx)
                dst_cell = ws.cell(row, col_idx)
                if src_cell.has_style:
                    dst_cell._style = copy(src_cell._style)
                if src_cell.number_format:
                    dst_cell.number_format = src_cell.number_format
                dst_cell.alignment = copy(src_cell.alignment)
                dst_cell.font = copy(src_cell.font)
                dst_cell.fill = copy(src_cell.fill)
                dst_cell.border = copy(src_cell.border)


def write_difference_sheet(wb: Workbook, settlement_ids: Set[str], shipment_ids: Set[str], abnormal_rows: Sequence[Sequence[Any]]) -> None:
    ws = wb.create_sheet("差异")
    headers = ["差异类型", "订单号", "说明"]
    ws.append(headers)
    for order_id in sorted(settlement_ids - shipment_ids):
        ws.append(["结算有，发货表无", order_id, "该订单进入income结算，但未在5月发货表tk订单号中找到"])
    for order_id in sorted(shipment_ids - settlement_ids):
        ws.append(["发货表有，结算无", order_id, "该订单在5月发货表中，但未进入本次KQ income结算"])
    for row in abnormal_rows:
        ws.append(list(row))
    if ws.max_row == 1:
        ws.append(["无差异", "", "结算订单与发货表订单一致"])
    style_raw_sheet(ws, 3)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 60


def write_product_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("商品")
    ws.append(["订单号", "店铺", "商品名称", "商品成本RMB", "备注"])
    style_raw_sheet(ws, 5)
    for col, width in {"A": 24, "B": 20, "C": 48, "D": 16, "E": 36}.items():
        ws.column_dimensions[col].width = width


def build_settlement() -> Path:
    script_dir = Path(__file__).resolve().parent
    data_root = script_dir.parents[2] / "data" / "航海计划结算" / "杭州克勤"
    data_root.mkdir(parents=True, exist_ok=True)
    base_dir = find_month_dir(data_root)
    refund_order_ids, refund_reasons = load_refund_orders(base_dir)
    stores = [load_store_data(files, refund_order_ids, refund_reasons) for files in pair_store_files(base_dir)]
    settlement_ids = set().union(*(item.totals.order_ids for item in stores)) if stores else set()
    abnormal_rows = [row for item in stores for row in item.abnormal_rows]
    jd_file = find_jd_file(base_dir)
    shipment_file = find_shipment_file(base_dir)
    shipment_sheet_name, shipment_headers, shipment_rows, shipment_ids = load_shipment_rows(shipment_file)
    jd_headers, jd_rows, packages = load_jd_rows(jd_file, settlement_ids)
    logistics = allocated_logistics(packages, settlement_ids)

    wb = Workbook()
    wb.remove(wb.active)
    write_summary(wb, stores, logistics, base_dir, jd_file, shipment_file)
    for item in stores:
        write_raw_sheet(wb, f"{item.totals.store} income", item.income_headers, item.income_rows)
        write_raw_sheet(wb, f"{item.totals.store} All Orders", item.order_headers, item.order_rows)
    write_raw_sheet(wb, "物流", jd_headers, jd_rows)
    write_raw_sheet(wb, "发货表", shipment_headers, shipment_rows)
    write_difference_sheet(wb, settlement_ids, shipment_ids, abnormal_rows)
    write_product_sheet(wb)
    apply_format_template(wb, script_dir)

    output = base_dir / f"KQ_TikTok_结算单_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
    wb.save(output)
    return output


def main() -> int:
    output = build_settlement()
    print("KQ结算单生成完成")
    print(f"输出文件：{output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
