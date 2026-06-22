#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
TikTok 订单 -> 模板-发货单-V2.xlsx

日常用法：
1. 把本脚本、模板-发货单-V2.xlsx、tiktok_order_list.xlsx 放在同一个目录。
2. 运行：python tiktok_shipping_list_generator.py
3. 固定生成：tiktok_order_list_发货单.xlsx
4. 固定日志：tiktok_shipping_list_generator.log
"""

from __future__ import annotations

import argparse
import logging
import sys
from collections import OrderedDict
from copy import copy
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill


SCRIPT_VERSION = "tiktok_shipping_list_generator_v20260615_v2_template"

DEFAULT_INPUT_NAME = "tiktok_order_list.xlsx"
DEFAULT_OUTPUT_NAME = "tiktok_order_list_发货单.xlsx"
DEFAULT_TEMPLATE_NAME = "模板-发货单-V2.xlsx"
DEFAULT_LOG_NAME = "tiktok_shipping_list_generator.log"
DEFAULT_CHANNEL = "VLG-TK"
DEFAULT_SHOP_NAME = "VLG TCG"
DEFAULT_MAX_ORDERS_PER_PACKAGE = 20


def default_data_dir(script_dir: Path) -> Path:
    return script_dir.parents[1] / "data" / "分拣"
AWAITING_SHIPMENT_TEXT = "awaiting shipment"
MASKED_ROW_FILL = PatternFill(fill_type="solid", fgColor="FFC7CE")


@dataclass
class OrderRow:
    order_id: str
    product_name: str
    order_amount: Decimal
    quantity: Decimal
    buyer_nickname: str
    zipcode: str
    recipient: str
    phone: str
    city_state: str
    street_address: str
    has_masked_privacy: bool = False


@dataclass
class ShippingRow:
    order_ids: List[str]
    product_names: List[str]
    amount: Decimal
    quantity: Decimal
    buyer_nicknames: List[str]
    zipcode: str
    recipient: str
    phone: str
    city_state: str
    street_address: str
    has_masked_privacy: bool = False


def clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip().rstrip("\t").strip()


def normalize_header(value: Any) -> str:
    return "".join(clean(value).split()).lower()


def to_decimal(value: Any) -> Decimal:
    text = clean(value).replace(",", "").replace("$", "")
    if not text:
        return Decimal("0")
    try:
        return Decimal(text)
    except InvalidOperation:
        return Decimal("0")


def display_number(value: Decimal) -> Any:
    if value == value.to_integral_value():
        return int(value)
    return float(value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


def unique_keep_order(values: Iterable[str]) -> List[str]:
    result: List[str] = []
    seen = set()
    for value in values:
        text = clean(value)
        if text and text not in seen:
            result.append(text)
            seen.add(text)
    return result


def join_nonempty(parts: Iterable[Any], separator: str = " ") -> str:
    return separator.join(clean(part) for part in parts if clean(part))


def contains_masked_privacy(*values: Any) -> bool:
    return any("*" in clean(value) for value in values)


def setup_logging(base_dir: Path) -> Path:
    log_file = base_dir / DEFAULT_LOG_NAME
    logging.basicConfig(
        filename=log_file,
        filemode="w",
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        encoding="utf-8",
        force=True,
    )
    return log_file


def resolve_input_path(input_arg: Optional[str], script_dir: Path, data_dir: Path) -> Path:
    if input_arg:
        input_file = Path(input_arg)
        if not input_file.is_absolute():
            input_file = Path.cwd() / input_file
        return input_file
    return data_dir / DEFAULT_INPUT_NAME


def resolve_template_path(input_file: Path, template_arg: Optional[str], script_dir: Path) -> Path:
    if template_arg:
        template_file = Path(template_arg)
        if not template_file.is_absolute():
            template_file = Path.cwd() / template_file
        return template_file

    same_dir_template = input_file.parent / DEFAULT_TEMPLATE_NAME
    if same_dir_template.exists():
        return same_dir_template
    return script_dir / DEFAULT_TEMPLATE_NAME


def resolve_output_path(input_file: Path, output_arg: Optional[str]) -> Path:
    if output_arg:
        output_file = Path(output_arg)
        if not output_file.is_absolute():
            output_file = Path.cwd() / output_file
        return output_file
    return input_file.parent / DEFAULT_OUTPUT_NAME


def normalize_sale_date(date_text: Optional[str]) -> str:
    if not date_text:
        return datetime.now().strftime("%Y%m%d")

    text = clean(date_text)
    if len(text) == 8 and text.isdigit():
        return text
    if len(text) == 4 and text.isdigit():
        return f"{datetime.now().year}{text}"
    raise ValueError("--date 必须是 YYYYMMDD 或 MMDD 格式，例如 20260615 或 0615")


def source_column_map(ws) -> Dict[str, int]:
    mapping: Dict[str, int] = {}
    for idx, cell in enumerate(ws[1], start=1):
        key = normalize_header(cell.value)
        if key:
            mapping[key] = idx
    return mapping


def get_source_col(mapping: Dict[str, int], name: str, required: bool = True) -> Optional[int]:
    col = mapping.get(normalize_header(name))
    if required and col is None:
        raise KeyError(f"源数据缺少必要字段：{name}")
    return col


def get_first_source_col(mapping: Dict[str, int], names: Sequence[str], required: bool = True) -> Optional[int]:
    for name in names:
        col = mapping.get(normalize_header(name))
        if col is not None:
            return col
    if required:
        raise KeyError(f"源数据缺少必要字段：{' / '.join(names)}")
    return None


def cell_value(row: Sequence[Any], col: Optional[int]) -> Any:
    if col is None:
        return None
    idx = col - 1
    if idx >= len(row):
        return None
    return row[idx]


def read_order_rows(input_file: Path) -> Tuple[List[OrderRow], int, int, int, int]:
    wb = load_workbook(input_file, data_only=True)
    ws = wb["OrderSKUList"] if "OrderSKUList" in wb.sheetnames else wb.active
    mapping = source_column_map(ws)

    required_names = [
        "Order ID",
        "Product Name",
        "Quantity",
        "Order Amount",
        "Buyer Nickname",
        "Zipcode",
        "Recipient",
        "Phone #",
        "State",
        "City",
        "Address Line 1",
        "Delivery Instruction",
        "Order Substatus",
    ]
    cols: Dict[str, Optional[int]] = {name: get_source_col(mapping, name, required=True) for name in required_names}
    cols["Address Line 2"] = get_first_source_col(mapping, ["Address Line 2", "Address Line"], required=False)
    cols["Shipping Information"] = get_source_col(mapping, "Shipping Information", required=False)

    rows: List[OrderRow] = []
    skipped_canceled = 0
    skipped_non_awaiting = 0
    skipped_zero_quantity = 0
    masked_awaiting_rows = 0

    for excel_row_num, raw_row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        order_id = clean(cell_value(raw_row, cols["Order ID"]))
        if not order_id or order_id.lower().startswith("platform unique order id"):
            continue

        order_substatus = clean(cell_value(raw_row, cols["Order Substatus"]))
        if order_substatus.lower() != AWAITING_SHIPMENT_TEXT:
            skipped_non_awaiting += 1
            if "cancel" in order_substatus.lower():
                skipped_canceled += 1
            logging.info("跳过第 %s 行，Order ID=%s，Order Substatus=%s", excel_row_num, order_id, order_substatus)
            continue

        quantity = to_decimal(cell_value(raw_row, cols["Quantity"]))
        if quantity <= 0:
            skipped_zero_quantity += 1
            logging.info("跳过第 %s 行，Order ID=%s，Quantity=%s", excel_row_num, order_id, quantity)
            continue

        state = clean(cell_value(raw_row, cols["State"]))
        city = clean(cell_value(raw_row, cols["City"]))
        line1 = clean(cell_value(raw_row, cols["Address Line 1"]))
        line2 = clean(cell_value(raw_row, cols["Address Line 2"]))
        instruction = clean(cell_value(raw_row, cols["Delivery Instruction"]))
        shipping_information = clean(cell_value(raw_row, cols["Shipping Information"]))
        city_state = join_nonempty([city, state], ", ")
        street_address = join_nonempty([line1, line2, instruction], " ")

        recipient = clean(cell_value(raw_row, cols["Recipient"]))
        phone = clean(cell_value(raw_row, cols["Phone #"]))
        zipcode = clean(cell_value(raw_row, cols["Zipcode"]))
        has_masked = contains_masked_privacy(
            recipient,
            phone,
            state,
            city,
            zipcode,
            line1,
            line2,
            instruction,
            shipping_information,
        )
        if has_masked:
            masked_awaiting_rows += 1

        rows.append(
            OrderRow(
                order_id=order_id,
                product_name=clean(cell_value(raw_row, cols["Product Name"])),
                order_amount=to_decimal(cell_value(raw_row, cols["Order Amount"])),
                quantity=quantity,
                buyer_nickname=clean(cell_value(raw_row, cols["Buyer Nickname"])),
                zipcode=zipcode,
                recipient=recipient,
                phone=phone,
                city_state=city_state,
                street_address=street_address,
                has_masked_privacy=has_masked,
            )
        )

    return rows, skipped_canceled, skipped_non_awaiting, skipped_zero_quantity, masked_awaiting_rows


def shipping_key(row: OrderRow) -> Tuple[str, ...]:
    if row.has_masked_privacy:
        return ("__MASKED_PRIVACY__", row.order_id)
    return (row.zipcode, row.recipient, row.phone, row.city_state, row.street_address)


def make_shipping_rows(rows: List[OrderRow], max_orders_per_package: int) -> List[ShippingRow]:
    if max_orders_per_package < 1:
        raise ValueError("max-orders 必须大于等于 1")

    groups: "OrderedDict[Tuple[str, ...], List[OrderRow]]" = OrderedDict()
    for row in rows:
        groups.setdefault(shipping_key(row), []).append(row)

    shipping_rows: List[ShippingRow] = []
    for group_rows in groups.values():
        order_ids_all = unique_keep_order(row.order_id for row in group_rows)
        for start in range(0, len(order_ids_all), max_orders_per_package):
            chunk_order_ids = order_ids_all[start : start + max_orders_per_package]
            chunk_order_id_set = set(chunk_order_ids)
            chunk_rows = [row for row in group_rows if row.order_id in chunk_order_id_set]
            if not chunk_rows:
                continue

            amount_by_order: "OrderedDict[str, Decimal]" = OrderedDict()
            for row in chunk_rows:
                if row.order_id not in amount_by_order:
                    amount_by_order[row.order_id] = row.order_amount

            first = chunk_rows[0]
            shipping_rows.append(
                ShippingRow(
                    order_ids=chunk_order_ids,
                    product_names=[row.product_name for row in chunk_rows if row.product_name],
                    amount=sum(amount_by_order.values(), Decimal("0")),
                    quantity=sum((row.quantity for row in chunk_rows), Decimal("0")),
                    buyer_nicknames=unique_keep_order(row.buyer_nickname for row in chunk_rows),
                    zipcode=first.zipcode,
                    recipient=first.recipient,
                    phone=first.phone,
                    city_state=first.city_state,
                    street_address=first.street_address,
                    has_masked_privacy=any(row.has_masked_privacy for row in chunk_rows),
                )
            )

    return shipping_rows


def target_column_map(ws) -> Dict[str, int]:
    mapping: Dict[str, int] = {}
    for cell in ws[1]:
        key = normalize_header(cell.value)
        if key:
            mapping[key] = cell.column
    return mapping


def find_target_col(tcols: Dict[str, int], aliases: Sequence[str], logical_name: str) -> int:
    normalized_aliases = [normalize_header(alias) for alias in aliases]
    for alias in normalized_aliases:
        col = tcols.get(alias)
        if col is not None:
            return col

    for header, col in tcols.items():
        if any(alias and (alias in header or header in alias) for alias in normalized_aliases):
            return col

    raise KeyError(f"模板缺少必要列：{logical_name}，可接受表头：{', '.join(aliases)}")


def copy_row_style(ws, source_row: int, target_row: int) -> None:
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for col_idx in range(1, ws.max_column + 1):
        src = ws.cell(source_row, col_idx)
        dst = ws.cell(target_row, col_idx)
        if src.has_style:
            dst._style = copy(src._style)
        dst.number_format = src.number_format
        dst.alignment = copy(src.alignment)
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.protection = copy(src.protection)


def clear_data_area(ws, last_row: int) -> None:
    for row_idx in range(2, last_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row_idx, col_idx)
            cell.value = None
            cell.hyperlink = None
            cell.comment = None


def write_text_cell(cell, value: Any) -> None:
    cell.number_format = "@"
    cell.value = clean(value)
    cell.data_type = "s"


def write_shipping_file(
    template_file: Path,
    output_file: Path,
    shipping_rows: List[ShippingRow],
    sale_date: str,
    channel: str,
    shop_name: str,
) -> None:
    wb = load_workbook(template_file)
    ws = wb["卡牌发货包导入"] if "卡牌发货包导入" in wb.sheetnames else wb.active
    tcols = target_column_map(ws)

    col = {
        "channel": find_target_col(tcols, ["渠道", "Channel"], "渠道/Channel"),
        "sale_date": find_target_col(tcols, ["售出日期", "Sale Date"], "售出日期/Sale Date"),
        "remark": find_target_col(tcols, ["备注", "Remark"], "备注/Remark"),
        "shop_name": find_target_col(tcols, ["店铺名称", "Shop Name"], "店铺名称/Shop Name"),
        "order_no": find_target_col(tcols, ["订单编号", "Order No.", "Order No"], "订单编号/Order No."),
        "goods_name": find_target_col(tcols, ["商品名称", "Goods Name"], "商品名称/Goods Name"),
        "total_price": find_target_col(tcols, ["总价", "Total Price"], "总价/Total Price"),
        "quantity": find_target_col(tcols, ["数量", "Qty"], "数量/Qty"),
        "nickname": find_target_col(tcols, ["昵称", "Nickname"], "昵称/Nickname"),
        "zipcode": find_target_col(tcols, ["邮编", "Zip Code", "Zipcode"], "邮编/Zip Code"),
        "receiver": find_target_col(tcols, ["收件人", "Receiver"], "收件人/Receiver"),
        "phone": find_target_col(tcols, ["收件电话", "Phone"], "收件电话/Phone"),
        "city_state": find_target_col(tcols, ["城市,州", "City, State"], "城市,州/City, State"),
        "street_address": find_target_col(tcols, ["街道地址", "Street Address"], "街道地址/Street Address"),
    }

    needed_last_row = max(2, len(shipping_rows) + 1)
    existing_last_row = max(ws.max_row, needed_last_row)

    while ws.max_row < needed_last_row:
        ws.append([None] * ws.max_column)

    for row_idx in range(2, needed_last_row + 1):
        copy_row_style(ws, 2, row_idx)

    clear_data_area(ws, max(existing_last_row, ws.max_row))

    for row_idx, shipping_row in enumerate(shipping_rows, start=2):
        ws.cell(row_idx, col["channel"]).value = channel
        write_text_cell(ws.cell(row_idx, col["sale_date"]), sale_date)
        ws.cell(row_idx, col["remark"]).value = None
        ws.cell(row_idx, col["shop_name"]).value = shop_name
        ws.cell(row_idx, col["order_no"]).value = "\n".join(shipping_row.order_ids)
        ws.cell(row_idx, col["goods_name"]).value = "\n".join(shipping_row.product_names)
        ws.cell(row_idx, col["total_price"]).value = display_number(shipping_row.amount)
        ws.cell(row_idx, col["quantity"]).value = display_number(shipping_row.quantity)
        ws.cell(row_idx, col["nickname"]).value = "\n".join(shipping_row.buyer_nicknames)
        write_text_cell(ws.cell(row_idx, col["zipcode"]), shipping_row.zipcode)
        ws.cell(row_idx, col["receiver"]).value = shipping_row.recipient
        ws.cell(row_idx, col["phone"]).value = shipping_row.phone
        ws.cell(row_idx, col["city_state"]).value = shipping_row.city_state
        ws.cell(row_idx, col["street_address"]).value = shipping_row.street_address

        for key in ["order_no", "goods_name", "nickname", "city_state", "street_address"]:
            cell = ws.cell(row_idx, col[key])
            cell.alignment = Alignment(
                horizontal=cell.alignment.horizontal,
                vertical=cell.alignment.vertical or "center",
                wrap_text=True,
            )

        if shipping_row.has_masked_privacy:
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row_idx, col_idx).fill = copy(MASKED_ROW_FILL)

    wb.save(output_file)


def verify_output(output_file: Path, expected_rows: int) -> None:
    wb = load_workbook(output_file, data_only=True)
    ws = wb["卡牌发货包导入"] if "卡牌发货包导入" in wb.sheetnames else wb.active
    if expected_rows <= 0:
        return
    first_order_no = clean(ws.cell(2, 5).value)
    if not first_order_no:
        raise RuntimeError("输出文件自检失败：第一条订单编号为空")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="TikTok 订单 Excel 转发货单 V2 模板")
    parser.add_argument("input", nargs="?", help=f"订单文件；不填默认读取同目录 {DEFAULT_INPUT_NAME}")
    parser.add_argument("--template", help=f"模板文件；默认优先使用订单同目录的 {DEFAULT_TEMPLATE_NAME}")
    parser.add_argument("--output", help=f"输出文件；默认生成 {DEFAULT_OUTPUT_NAME}")
    parser.add_argument("--date", help="售出日期，格式 YYYYMMDD 或 MMDD；不填使用运行当天日期")
    parser.add_argument("--channel", default=DEFAULT_CHANNEL, help=f"渠道，默认 {DEFAULT_CHANNEL}")
    parser.add_argument("--shop-name", default=DEFAULT_SHOP_NAME, help=f"店铺名称，默认 {DEFAULT_SHOP_NAME}")
    parser.add_argument("--max-orders", type=int, default=DEFAULT_MAX_ORDERS_PER_PACKAGE, help="每行最多合并多少个唯一订单编号，默认 20")
    return parser.parse_args()


def run() -> None:
    args = parse_args()
    script_dir = Path(__file__).resolve().parent
    data_dir = default_data_dir(script_dir)
    data_dir.mkdir(parents=True, exist_ok=True)
    log_file = setup_logging(data_dir)

    input_file = resolve_input_path(args.input, script_dir, data_dir)
    if input_file.suffix.lower() != ".xlsx":
        raise ValueError("目前脚本只支持 .xlsx 文件；如果是 .xls，请先另存为 .xlsx。")
    if not input_file.exists():
        raise FileNotFoundError(f"订单文件不存在：{input_file}。请将订单文件命名为 {DEFAULT_INPUT_NAME} 并放到脚本同目录。")

    template_file = resolve_template_path(input_file, args.template, script_dir)
    if not template_file.exists():
        raise FileNotFoundError(f"模板文件不存在：{template_file}")

    sale_date = normalize_sale_date(args.date)
    output_file = resolve_output_path(input_file, args.output)

    logging.info("脚本版本：%s", SCRIPT_VERSION)
    logging.info("订单文件：%s", input_file)
    logging.info("模板文件：%s", template_file)
    logging.info("输出文件：%s", output_file)
    logging.info("售出日期：%s", sale_date)
    logging.info("渠道：%s", args.channel)
    logging.info("店铺名称：%s", args.shop_name)

    rows, skipped_canceled, skipped_non_awaiting, skipped_zero_quantity, masked_awaiting_rows = read_order_rows(input_file)
    shipping_rows = make_shipping_rows(rows, max_orders_per_package=args.max_orders)
    write_shipping_file(
        template_file=template_file,
        output_file=output_file,
        shipping_rows=shipping_rows,
        sale_date=sale_date,
        channel=args.channel,
        shop_name=args.shop_name,
    )
    verify_output(output_file, expected_rows=len(shipping_rows))

    logging.info("处理完成")
    logging.info("有效 SKU 行：%s", len(rows))
    logging.info("跳过取消 SKU 行：%s", skipped_canceled)
    logging.info("跳过非 Awaiting shipment SKU 行：%s", skipped_non_awaiting)
    logging.info("跳过 Quantity 小于等于 0 SKU 行：%s", skipped_zero_quantity)
    logging.info("含星号待处理 SKU 行：%s", masked_awaiting_rows)
    logging.info("生成发货行：%s", len(shipping_rows))

    print("处理完成")
    print(f"脚本版本：{SCRIPT_VERSION}")
    print(f"订单文件：{input_file}")
    print(f"模板文件：{template_file}")
    print(f"输出文件：{output_file}")
    print(f"日志文件：{log_file}")
    print(f"售出日期：{sale_date}")
    print(f"渠道：{args.channel}")
    print(f"店铺名称：{args.shop_name}")
    print(f"有效 SKU 行：{len(rows)}")
    print(f"跳过取消 SKU 行：{skipped_canceled}")
    print(f"跳过非 Awaiting shipment SKU 行：{skipped_non_awaiting}")
    print(f"跳过 Quantity 小于等于 0 SKU 行：{skipped_zero_quantity}")
    print(f"含星号待处理 SKU 行：{masked_awaiting_rows}")
    print(f"生成发货行：{len(shipping_rows)}")


def main() -> int:
    script_dir = Path(__file__).resolve().parent
    data_dir = default_data_dir(script_dir)
    data_dir.mkdir(parents=True, exist_ok=True)
    log_file = setup_logging(data_dir)
    try:
        run()
        return 0
    except Exception as exc:
        logging.exception("程序运行失败")
        print("运行失败，错误已写入日志文件：")
        print(log_file)
        print(f"错误信息：{exc}")
        return 1


if __name__ == "__main__":
    sys.exit(main())
